
from collections import defaultdict
from django.shortcuts import render
from django.contrib import messages
from django.http import HttpResponse, FileResponse, JsonResponse
from django.db.models import Q, Count, Case, When, IntegerField, Sum, Min
from django.conf import settings
from django.utils import timezone
from django.shortcuts import render
from .models import Usagereport, ReportGeneration, ENQUIRY_RATES, SubscriberProductRate
from datetime import date, timedelta, datetime
import calendar
import io
import os
import re
import zipfile
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import os.path
import uuid
from django.urls import reverse
from django.contrib.auth.decorators import login_required,user_passes_test
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import logging
from django.core.cache import cache
import hashlib
from django.db.models import DateField
from django.db.models.functions import Cast
from django.db.models import Exists, OuterRef
import csv

# Get an instance of a logger
logger = logging.getLogger(__name__)


# Add this list near the top of your views.py
KEY_SUBSCRIBERS = [
    "Addosser MFB",
    "ALTRACRED FINANCE INVESTMENT LIMITED",
    "AppZone Limited Lagos",
    "AutoChek FInancing SPV Limited",
    "Boctrust MFB",
    "Branch International Financial Services Limited",
    "Bravewood Finance Company Limited",
    "Cashigo International Limited",
    "CashXpress",
    "CREDITCHECK AFRICA",
    "Crednet Technologies Limited",
    "ECOBANK Nigeria PLC Lagos",
    "Earnipay",
    "Fidelity Bank Plc Lagos",
    "Fina Trust Microfinance Bank Limited",
    "First Bank Plc Lagos",
    "First City Monument Bank Limited Lagos",
    "FLOURISH MFB",
    "Infinity MFB",
    "Keystone Bank",
    "KOBOGO NIGERIA LIMITED",
    "KWIKPAY",
    "LAPO Microfinance Bank Limited Edo",
    "LINKS MICROFINANCE BANK LTD",
    "MAINLINE DIGITECH  INNOVATIONS LIMITED",
    "Mutual Trust Microfinance Bank Ltd Abuja",
    "NewEdge Finance Limited Lagos",
    "NPF Microfinance Bank PLC Lagos",
    "Polaris Bank Limited",
    "RIMWORLD CAPITAL",
    "Rosabon Finance Lagos",
    "Stanbic IBTC Bank PLC",
    "UCEE Microfinance Bank Limited",
    "Unity Bank",
    "Wema Bank",
    "Pebblescore Data Service Limited",
    "Lending Technologies Limited(Lendsqr)"
]

# Function to safely populate direct assignments by checking for merged cells first
def safe_cell_assignment(ws, row, col, value):
    """Helper function to safely assign values to cells, handling merged cells."""
    write_to_cell(ws, row, col, value)

def write_to_cell(ws, row, col, value):
    """
    Safely writes a value to a cell, handling merged cells and preserving formatting.
    
    When using the direct cell assignment approach like in VBA, we need to 
    handle merged cells specially as they are read-only in openpyxl.
    """
    coordinate = ws.cell(row=row, column=col).coordinate
    is_merged = False
    merged_range_to_restore = None
    
    # Store the original cell style before unmerging
    original_cell = ws.cell(row=row, column=col)
    original_style = original_cell._style
    original_number_format = original_cell.number_format
    
    # Special handling for product name header (row 32)
    if row == 32 and col == 4:  # D32
        # Try additional columns for row 32 as it might be a merged cell
        for try_col in range(1, 10):  # Try columns A through I
            try:
                ws.cell(row=row, column=try_col).value = value
            except:
                pass
    
    # Check if the cell is in a merged range
    for merged_range in list(ws.merged_cells.ranges):
        if coordinate in merged_range:
            is_merged = True
            merged_range_to_restore = str(merged_range)
            ws.unmerge_cells(merged_range_to_restore)
            break
    
    # Now we can safely set the value
    target_cell = ws.cell(row=row, column=col)
    target_cell.value = value
    
    # Restore the original style
    target_cell._style = original_style
    if original_number_format != 'General':
        target_cell.number_format = original_number_format
    
    # Restore the merge if needed
    if is_merged and merged_range_to_restore:
        ws.merge_cells(merged_range_to_restore)

# Function to safely populate direct assignments by checking for merged cells first
def safe_cell_assignment(ws, row, col, value):
    """Helper function to safely assign values to cells, handling merged cells."""
    write_to_cell(ws, row, col, value)

@login_required
def home(request):
    """Home page view with subscriber list and date range."""
    today = date.today()
    first_day_of_month = today.replace(day=1)
    
    # Calculate first day of next month
    if today.month == 12:
        first_day_next_month = date(today.year + 1, 1, 1)
    else:
        first_day_next_month = date(today.year, today.month + 1, 1)
    
    # Use Django ORM to fetch distinct subscriber names
    start_date = first_day_of_month
    end_date = first_day_next_month
    
    subscribers = Usagereport.objects.filter(
        DetailsViewedDate__gte=start_date,
        DetailsViewedDate__lt=end_date
    ).values_list('SubscriberName', flat=True).distinct().order_by('SubscriberName')

    # Format dates as strings for the template
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')

    context = {
        'subscribers': subscribers,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }
    return render(request, 'bulkrep/home.html', context)

def clean_filename(filename):
    """Clean subscriber name for a valid filename, similar to VBA script."""
    # Replace invalid characters with hyphens
    invalid_chars = r'[\/\\\:\*\?"<>\|]'
    return re.sub(invalid_chars, '-', filename)

@login_required
def single_report(request):
    """View for generating a single report."""
    # --- Initial Setup (Largely Unchanged) ---
    today = date.today()
    first_day_of_month = today.replace(day=1)

    if today.month == 12:
        first_day_next_month = date(today.year + 1, 1, 1)
    else:
        first_day_next_month = date(today.year, today.month + 1, 1)

    report_gen = None
    subscribers = Usagereport.objects.values_list('SubscriberName', flat=True).distinct().order_by('SubscriberName')

    context = {
        'subscribers': subscribers,
        'start_date': first_day_of_month.strftime('%Y-%m-%d'),
        'end_date': first_day_next_month.strftime('%Y-%m-%d'),
    }

    if request.method == 'POST':
        # --- Form Data & Report Tracking (Unchanged) ---
        subscriber_id = request.POST.get('subscriber_id')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        include_bills = request.POST.get('include_bills') == 'on'
        include_products = request.POST.get('include_products') == 'on'

        if subscriber_id:
            report_gen = ReportGeneration.objects.create(
                user=request.user, generator=request.user.username, report_type='single',
                status='in_progress', subscriber_name=subscriber_id,
                from_date=start_date_str, to_date=end_date_str
            )

        if not subscriber_id:
            messages.error(request, "Please select a subscriber.")
            return render(request, 'bulkrep/single_report.html', context)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            start_date_display = start_date.strftime('%d/%m/%Y')
            end_date_display = end_date.strftime('%d/%m/%Y')
        except (ValueError, TypeError) as e:
            messages.error(request, f"Invalid date format: {str(e)}")
            return render(request, 'bulkrep/single_report.html', context)

        # --- OPTIMIZED: Fetch all product counts in a single query ---
        summary_bills = {}
        if include_bills:
            queryset = Usagereport.objects.filter(
                DetailsViewedDate__gte=start_date,
                DetailsViewedDate__lte=end_date,
                SubscriberName=subscriber_id
            )
            summary_bills = queryset.aggregate(
                consumer_snap_check=Count(Case(When(ProductName__icontains='Snap Check', then=1))),
                consumer_basic_trace=Count(Case(When(ProductName__icontains='Basic Trace', then=1))),
                consumer_basic_credit=Count(Case(When(ProductName__icontains='Basic Credit', then=1))),
                consumer_detailed_credit=Count(Case(When(ProductName__icontains='Detailed Credit',then=1))),
                xscore_consumer_detailed_credit=Count(Case(When(ProductName__icontains='X-SCore Consumer Detailed Credit', then=1))),
                commercial_basic_trace=Count(Case(When(ProductName__icontains='Commercial Basic Trace', then=1))),
                commercial_detailed_credit=Count(Case(When(ProductName__icontains='Commercial detailed Credit', then=1))),
                enquiry_report=Count(Case(When(ProductName__icontains='Enquiry Report', then=1))),
                consumer_dud_cheque=Count(Case(When(ProductName__icontains='Consumer Dud Cheque', then=1))),
                commercial_dud_cheque=Count(Case(When(ProductName__icontains='Commercial Dud Cheque', then=1))),
                director_basic_report=Count(Case(When(ProductName__icontains='Director Basic Report', then=1))),
                director_detailed_report=Count(Case(When(ProductName__icontains='Director Detailed Report', then=1))),
            )


        # Query for product details using Django ORM
        if include_products:
            product_data = Usagereport.objects.filter(
                DetailsViewedDate__gte=start_date,
                DetailsViewedDate__lte=end_date,
                SubscriberName=subscriber_id
            ).order_by('ProductName', 'DetailsViewedDate').values(
                'SubscriberName', 'SystemUser', 'SearchIdentity', 'SubscriberEnquiryDate',
                'SearchOutput', 'DetailsViewedDate', 'ProductInputed', 'ProductName'
            )
            
            # Group by ProductName
            product_sections = {}
            for record in product_data:
                product_name = record['ProductName']
                if product_name not in product_sections:
                    product_sections[product_name] = []
                product_sections[product_name].append(record)
            
            product_data = list(product_data)  # Convert to list for compatibility
        else:
            product_sections = {}
            product_data = []

        if not product_data and include_products:
            messages.warning(request, f"No data found for subscriber {subscriber_id} between {start_date_display} and {end_date_display}.")
            return render(request, 'bulkrep/single_report.html', context)

        # Generate Excel report
        try:
            template_path = os.path.join(settings.MEDIA_ROOT, 'Templateuse.xlsx')
            buffer = io.BytesIO()
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
        except Exception as e:
            logger.error(f"Error loading Excel template: {str(e)}")
            messages.error(request, f"Error loading Excel template: {str(e)}")
            return render(request, 'bulkrep/single_report.html', context)
        try:
            # Look for "Productname" cell to identify where to put the dynamic product name
            product_name_cell = None
            for row in range(30, 35):  # Search rows 30-34
                for col in range(1, 10):  # Search columns A-I
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and "product" in str(cell_value).lower():
                        try:
                            product_name_cell = (row, col)
                        except Exception as e:
                            logger.error(f"Error finding product name cell: {str(e)}")
                            messages.error(request, f"Error finding product name cell: {str(e)}")
                            return render(request, 'bulkrep/single_report.html', context)
                        break
                if product_name_cell:
                    break
            
            # Write subscriber info directly to cells using the safe method
            # For row 2, which is merged from H to P, we need to set the value to the first cell in the merged range
            # First, find the merged range that contains row 2
            row2_merged_range = None
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == 2 and merged_range.max_row == 2:
                    row2_merged_range = merged_range
                    break
            
            # If we found a merged range for row 2, unmerge it, set the value with the required format, and remerge it
            if row2_merged_range:
                merged_range_str = str(row2_merged_range)
                ws.unmerge_cells(merged_range_str)
                # Set the value to the first cell in the merged range with the required format
                original_cell = ws.cell(row=2, column=row2_merged_range.min_col)
                new_content = f"FirstCentral NIGERIA - BILLING DETAILS - {subscriber_id}"
                original_cell.value = new_content
                
                # Remerge the cells
                ws.merge_cells(merged_range_str)
            else:
                # Fallback to the original method if no merged range is found
                safe_cell_assignment(ws, 2, 5, subscriber_id)  # E2
            
            safe_cell_assignment(ws, 5, 4, f"BILLING DETAILS - {subscriber_id}")  # D5
            
            # For row 6, which is merged from B to Q, we need to set the value to the first cell in the merged range
            # First, find the merged range that contains row 6
            row6_merged_range = None
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == 6 and merged_range.max_row == 6:
                    row6_merged_range = merged_range
                    break
            
            # If we found a merged range for row 6, unmerge it, set the value, and remerge it
            date_range_text = f"REPORT GENERATED FOR RECORDS BETWEEN {start_date_display} and {end_date_display}"
            if row6_merged_range:
                merged_range_str = str(row6_merged_range)
                ws.unmerge_cells(merged_range_str)
                # Set the value to the first cell in the merged range (column B = 2)
                ws.cell(row=6, column=2).value = date_range_text
                # Remerge the cells
                ws.merge_cells(merged_range_str)
            else:
                # Fallback to the original method if no merged range is found
                safe_cell_assignment(ws, 6, 4, date_range_text)  # D6
                
               
            
# --- OPTIMIZED: Billing Logic ---
            if include_bills:
                # Get all custom rates for the subscriber at once
                custom_rates_qs = SubscriberProductRate.objects.filter(subscriber_name__iexact=subscriber_id)
                custom_rates_lookup = {rate.product_name.lower(): rate.rate for rate in custom_rates_qs}

                def get_rate(product_name_key, product_name_display):
                    default_rate = ENQUIRY_RATES.get(product_name_key, Decimal('0.00'))
                    retrieved_rate = custom_rates_lookup.get(product_name_display.lower(), default_rate)
                    # return custom_rates_lookup.get(product_name_display.lower(), default_rate)
                    try:
                        return Decimal(retrieved_rate)
                    except (ValueError, TypeError, InvalidOperation):
                        # If conversion fails (e.g., rate is "N/A"), fall back to the default.
                        return default_rate

                # Use in-memory data instead of repeated DB queries
                total_amount = Decimal('0.00')
                products_to_bill = {
                    12: ('consumer_snap_check', 'Consumer Snap Check'), 13: ('consumer_basic_trace', 'Consumer Basic Trace'),
                    14: ('consumer_basic_credit', 'Consumer Basic Credit'), 15: ('consumer_detailed_credit', 'Consumer Detailed Credit'),
                    16: ('xscore_consumer_detailed_credit', 'X-Score Consumer Detailed Credit'), 17: ('commercial_basic_trace', 'Commercial Basic Trace'),
                    18: ('commercial_detailed_credit', 'Commercial Detailed Credit'), 20: ('enquiry_report', 'Enquiry Report'),
                    22: ('consumer_dud_cheque', 'Consumer Dud Cheque'), 23: ('commercial_dud_cheque', 'Commercial Dud Cheque'),
                    25: ('director_basic_report', 'Director Basic Report'), 26: ('director_detailed_report', 'Director Detailed Report'),
                }

                for row, (key, name) in products_to_bill.items():
                    quantity = summary_bills.get(key, 0)
                    rate = get_rate(key, name)
                    amount = Decimal(quantity) * rate
                    
                    safe_cell_assignment(ws, row, 9, quantity)
                    safe_cell_assignment(ws, row, 13, f"₦{rate:,.2f}")
                    safe_cell_assignment(ws, row, 16, f"₦{amount:,.2f}")
                    total_amount += amount

                vat_amount = total_amount * Decimal('0.075')
                amount_due = total_amount + vat_amount
                safe_cell_assignment(ws, 28, 16, f"₦{total_amount:,.2f}")
                safe_cell_assignment(ws, 29, 16, f"₦{vat_amount:,.2f}")
                safe_cell_assignment(ws, 30, 16, f"₦{amount_due:,.2f}")
                

            
            if include_products:
                start_row_offset = 36  # Initial start row for product sections on Sheet1
                current_sheet = ws
                sheet2 = wb["Sheet2"] if "Sheet2" in wb.sheetnames else None
                
                # Sort product_sections by product_name
                sorted_product_names = sorted(product_sections.keys())
                
                # First, find the original product name cell in the template (rows 32-35)
                product_name_cell = None
                for row in range(32, 36):  # Rows 32-35 (inclusive)
                    for col in range(1, 16):  # Assuming columns A-O are important
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and "product" in str(cell_value).lower():
                            product_name_cell = (row, col)
                            break
                    if product_name_cell:
                        break
                        
                # Save the template header structure (rows 32-35) for subsequent products
                header_template = []
                header_rows = (32, 35)  # Range of rows to copy for the header template
                for row in range(header_rows[0], header_rows[1] + 1):
                    row_data = []
                    for col in range(1, 16):  # Assuming columns A-O are important
                        cell = ws.cell(row=row, column=col)
                        # Store cell value and position only - we'll copy styles directly later
                        cell_info = {
                            'value': cell.value,
                            'position': (row, col)
                        }
                        # Check if cell is part of merged range
                        for m_range in ws.merged_cells.ranges:
                            if (row, col) == (m_range.min_row, m_range.min_col):
                                cell_info['merged'] = (m_range.max_row - m_range.min_row + 1, 
                                                      m_range.max_col - m_range.min_col + 1)
                                break
                        row_data.append(cell_info)
                    header_template.append(row_data)
                
                # Set data row start - will be used for first product                
                data_start_row = 36  # Initial start for data rows after template header
                lastProduct = ""
                
                # For each product, use appropriate header section
                for product_idx, product_name in enumerate(sorted_product_names):
                    product_records = product_sections[product_name]
                    
                    if product_idx == 0:
                        # For first product, use the existing template header
                        if product_name_cell:
                            row, col = product_name_cell
                            # Replace "Product Name" with actual first product name in template
                            safe_cell_assignment(ws, row, col, product_name)
                        current_sheet = ws
                        current_row_offset = data_start_row  # Start data at row 36
                        serial_number_base = data_start_row - 1
                    else:
                        # Add space between different products (extra spacing)
                        current_row_offset += 4  # Add significant spacing between products
                            
                        # Create new header for subsequent products
                        header_start_row = current_row_offset
                        
                        # Clone the header section for this product
                        for template_row_idx, template_row in enumerate(header_template):
                            target_row = header_start_row + template_row_idx
                        
                            # Ensure we're not past row limits
                            if current_sheet == ws and target_row > 1000000 and sheet2:
                                current_sheet = sheet2
                                # Reset for Sheet2
                                header_start_row = 13
                                target_row = header_start_row + template_row_idx
                        
                            # Unmerge any existing merged cells in the target area
                            for m_range in list(current_sheet.merged_cells.ranges):
                                if m_range.min_row <= target_row <= m_range.max_row:
                                    current_sheet.unmerge_cells(
                                        start_row=m_range.min_row, 
                                        start_column=m_range.min_col,
                                        end_row=m_range.max_row, 
                                        end_column=m_range.max_col
                                    )
                        
                            # Copy each cell from the template to the target area
                            for col_idx, cell_info in enumerate(template_row):
                                target_col = col_idx + 1
                                target_cell = current_sheet.cell(row=target_row, column=target_col)
                                
                                # Copy styles directly from the original cell
                                original_row, original_col = cell_info['position']
                                original_cell = ws.cell(row=original_row, column=original_col)
                                
                                # Copy cell format using openpyxl's built-in method
                                target_cell._style = copy(original_cell._style)
                                
                                # Set value using safe_cell_assignment to handle merged cells properly
                                if template_row_idx == product_name_cell[0] - header_rows[0] and \
                                   col_idx == product_name_cell[1] - 1:
                                    # For the product name cell, use the safe assignment method
                                    safe_cell_assignment(current_sheet, target_row, target_col, product_name)
                                elif cell_info['value'] is not None:
                                    # For other cells with values, use the safe assignment method
                                    safe_cell_assignment(current_sheet, target_row, target_col, cell_info['value'])
                                
                                # Recreate merged cells
                                if 'merged' in cell_info:
                                    rows, cols = cell_info['merged']
                                    current_sheet.merge_cells(
                                        start_row=target_row, 
                                        start_column=target_col,
                                        end_row=target_row + rows - 1, 
                                        end_column=target_col + cols - 1
                                    )
                        
                        # Update data start row to be after this new header
                        current_row_offset = header_start_row + (header_rows[1] - header_rows[0] + 1)
                        serial_number_base = current_row_offset - 1
                        
                        # Add header for the data section
                        safe_cell_assignment(current_sheet, current_row_offset - 1, 4, "Unique Tracking Number")

                    # Process data records for this product
                    for record_idx, record in enumerate(product_records):
                        current_row = current_row_offset + record_idx
                        
                        # Switch to Sheet2 when reaching max row (like VBA)
                        if current_row > 1000000 and sheet2 and current_sheet != sheet2:
                            current_sheet = sheet2
                            current_row_offset = 13  # Reset row offset for Sheet2 as per VBA
                            serial_number_base = 12  # Reset serial number base for Sheet2
                            current_row = current_row_offset + record_idx  # Recalculate current_row for Sheet2
                        
                        # Copy row format from template data row
                        template_data_row = 36
                        max_col = 17  # Extend to include all columns that need formatting
                        copy_row_format(current_sheet, template_data_row, current_row, max_col)
                        
                        # Now assign values as before
                        safe_cell_assignment(current_sheet, current_row, 2, current_row - serial_number_base)  # Serial Number
                        safe_cell_assignment(current_sheet, current_row, 3, "")  # Branch ID
                        safe_cell_assignment(current_sheet, current_row, 4, "")  # Unique Tracking Number column left blank
                        safe_cell_assignment(current_sheet, current_row, 5, record['SubscriberName'])
                        safe_cell_assignment(current_sheet, current_row, 7, record['SystemUser'] if record['SystemUser'] else "")
                        # Format SubscriberEnquiryDate
                        subscriber_enquiry_date = record['SubscriberEnquiryDate']
                        if subscriber_enquiry_date:

                            if isinstance(subscriber_enquiry_date, str):
                                date_str = subscriber_enquiry_date
                            else:
                                # Convert to date object first to remove any time component
                                if hasattr(subscriber_enquiry_date, 'date'):
                                    date_only = subscriber_enquiry_date.date()
                                else:
                                    date_only = subscriber_enquiry_date
                                date_str = date_only.strftime('%Y-%m-%d')
                            # Set cell format to text BEFORE assignment to prevent Excel from adding time component
                            current_sheet.cell(row=current_row, column=10).number_format = '@'
                            safe_cell_assignment(current_sheet, current_row, 10, date_str)
                            # Ensure the format stays as text after assignment
                            current_sheet.cell(row=current_row, column=10).number_format = '@'
                        else:
                            safe_cell_assignment(current_sheet, current_row, 10, "")
                        safe_cell_assignment(current_sheet, current_row, 11, record['ProductName'])
                        # Format DetailsViewedDate to match SubscriberEnquiryDate format
                        details_viewed_date = record['DetailsViewedDate']
                        if details_viewed_date:
                            if isinstance(details_viewed_date, str):
                                date_str = details_viewed_date
                            else:
                                # Convert to date object first to remove any time component
                                if hasattr(details_viewed_date, 'date'):
                                    date_only = details_viewed_date.date()
                                else:
                                    date_only = details_viewed_date
                                date_str = date_only.strftime('%Y-%m-%d')
                            # Set cell format to text BEFORE assignment to prevent Excel from adding time component
                            current_sheet.cell(row=current_row, column=12).number_format = '@'
                            safe_cell_assignment(current_sheet, current_row, 12, date_str)
                            # Ensure the format stays as text after assignment
                            current_sheet.cell(row=current_row, column=12).number_format = '@'
                        else:
                            safe_cell_assignment(current_sheet, current_row, 12, "")
                        safe_cell_assignment(current_sheet, current_row, 15, record['SearchOutput'] if record['SearchOutput'] else "")
                        
                        # Apply merge and center formatting to this data row
                        merge_and_center_data_row(current_sheet, current_row)
                    
                    # Move offset to after this product's data for next product
                    current_row_offset += len(product_records)
                    
                    # Set lastProduct for the next iteration
                    lastProduct = product_name

            # Auto-size columns for better readability
            for sheet in wb.worksheets:
                auto_size_columns(sheet)

            # Add 'Generated by: <username>' to the first available merged O-Q row, or row 8 if none, with bold and italic formatting
            add_generated_by(ws, request.user.username, current_row_offset - 1)

            # Save workbook to buffer
            wb.save(buffer)
            buffer.seek(0)
            
            # Generate filename and create directory if it doesn't exist
            month_year = start_date.strftime('%B%Y')
            clean_subscriber = clean_filename(subscriber_id)
            filename = f"{clean_subscriber}_{month_year}_{uuid.uuid4().hex[:8]}.xlsx"
            single_reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'single')
            os.makedirs(single_reports_dir, exist_ok=True)
            file_path = os.path.join(single_reports_dir, filename)
            
            # Write buffer contents to file using getvalue() to get entire content regardless of position
            with open(file_path, 'wb') as f:
                f.write(buffer.getvalue())
            download_url = settings.MEDIA_URL + f'reports/single/{filename}'
            
            # Update report generation status to success
            if 'report_gen' in locals() and report_gen:
                report_gen.status = 'success'
                report_gen.completed_at = timezone.now()
                report_gen.save()
                
            return render(request, 'bulkrep/download_ready.html', {
                'download_url': download_url
            })
            
        except Exception as e:
            error_msg = f"Error generating report: {str(e)}"
            messages.error(request, error_msg)
            if 'report_gen' in locals() and report_gen:
                report_gen.status = 'failed'
                report_gen.error_message = error_msg[:500]  # Truncate to fit in the field
                report_gen.completed_at = timezone.now()
                report_gen.save()
            return render(request, 'bulkrep/single_report.html', context)

    return render(request, 'bulkrep/single_report.html', context)

# Helper function to auto-size columns for better readability
def auto_size_columns(worksheet, min_col=1, max_col=17):
    """
    Auto-size columns in the worksheet to fit their contents with reasonable widths.
    Applies modest adjustments to prevent columns from being too wide.
    Handles merged cells properly, especially for rows 2 and 6.
    """
    # Define minimum and maximum widths for specific columns
    min_widths = {
        # Set minimum widths to prevent columns from being too narrow
        5: 10,  # SubscriberName (E)
        6: 10,  # SubscriberName (F)
        7: 9,  # SystemUser (G)
        8: 10,  # SystemUser (H)
        9: 9,  # SystemUser (I)
        10: 15, # SubscriberEnquiryDate (J)
        11: 20, # ProductName (K)
        12: 5, # DetailsViewedDate (L)
        13: 10, # DetailsViewedDate (M)
        14: 5, # DetailsViewedDate (N)
        # Default minimum width for other columns
        'default': 10
    }


    
    # Define maximum widths for specific columns that tend to get too wide
    max_widths = {
        # DetailsViewedDate (columns L-N, 12-14)
        12: 5, 13: 10, 14: 5,
        # SearchOutput (columns O-Q, 15-17) - increased to allow more content to be visible
        15: 40, 16: 40, 17: 40,
        # Default max width for other columns
        'default': 30
    }
    
    # Store merged ranges for special handling
    merged_ranges = list(worksheet.merged_cells.ranges)
    
    # Special handling for rows 2 and 6 which have specific merged ranges
    row2_merged_range = None
    row6_merged_range = None
    
    for merged_range in merged_ranges:
        if merged_range.min_row == 2 and merged_range.max_row == 2:
            row2_merged_range = merged_range
        elif merged_range.min_row == 6 and merged_range.max_row == 6:
            row6_merged_range = merged_range
    
    for col_idx in range(min_col, max_col + 1):
        # Get the maximum content width in the column
        max_length = 0
        column = worksheet.column_dimensions[get_column_letter(col_idx)]
        
        # Skip columns that are part of the merged ranges for rows 2 and 6
        # This prevents these merged cells from affecting column widths
        skip_column = False
        if row2_merged_range and col_idx >= row2_merged_range.min_col and col_idx <= row2_merged_range.max_col:
            # For row 2 merged range, only consider the first column in the range
            if col_idx > row2_merged_range.min_col:
                skip_column = True
        
        if row6_merged_range and col_idx >= row6_merged_range.min_col and col_idx <= row6_merged_range.max_col:
            # For row 6 merged range, only consider the first column in the range
            if col_idx > row6_merged_range.min_col:
                skip_column = True
        
        if not skip_column:
            # Check all cells in this column
            for row_idx, row in enumerate(worksheet.rows, 1):
                # Skip rows 2 and 6 for all columns except the first column in their merged ranges
                if (row_idx == 2 and row2_merged_range and col_idx != row2_merged_range.min_col) or \
                   (row_idx == 6 and row6_merged_range and col_idx != row6_merged_range.min_col):
                    continue
                
                if len(row) >= col_idx:
                    cell = row[col_idx-1]  # 0-based index
                    if cell.value:
                        # Calculate the approximate width based on content
                        try:
                            cell_length = len(str(cell.value))
                            # Adjust for merged cells
                            is_in_merge = False
                            for merged_range in merged_ranges:
                                if cell.coordinate in merged_range:
                                    # Special handling for different types of merged cells
                                    merge_width = merged_range.max_col - merged_range.min_col + 1
                                    
                                    # For rows 2 and 6, handle specially
                                    if (row_idx == 2 and merged_range == row2_merged_range) or \
                                       (row_idx == 6 and merged_range == row6_merged_range):
                                        # For the first column in the merge, allocate more space
                                        if col_idx == merged_range.min_col:
                                            # Allocate more space to first column but not all
                                            cell_length = cell_length * 0.4  # 40% to first column
                                        else:
                                            # Distribute remaining 60% evenly among other columns
                                            cell_length = (cell_length * 0.6) / (merge_width - 1)
                                    
                                    # Special handling for SearchOutput columns (O-Q, 15-17)
                                    elif 15 <= merged_range.min_col <= 17 and 15 <= merged_range.max_col <= 17:
                                        # For SearchOutput with wrap_text, distribute more evenly
                                        # Give more width to first column for better readability
                                        if col_idx == merged_range.min_col:
                                            cell_length = cell_length * 0.5  # 50% to first column
                                        else:
                                            # Distribute remaining 50% evenly
                                            cell_length = (cell_length * 0.5) / (merge_width - 1)
                                    else:
                                        # For other merged cells, divide by the number of columns
                                        # but ensure a minimum reasonable length
                                        cell_length = max(10, cell_length / merge_width)
                                    
                                    is_in_merge = True
                                    break
                            
                            # Update max_length regardless of merge status
                            # This ensures all cells contribute to column width calculation
                            max_length = max(max_length, cell_length)
                        except:
                            pass
        
        # Set column width with appropriate padding and respect minimum/maximum widths
        if max_length > 0:
            # Add padding based on content type - more padding for text columns
            if col_idx >= 15 and col_idx <= 17:  # SearchOutput columns
                # For SearchOutput columns with wrap_text, use more generous padding
                calculated_width = max_length + 3
            else:
                # For other columns, use standard padding
                calculated_width = max_length + 2
            
            # Apply column-specific minimum and maximum width constraints
            column_min_width = min_widths.get(col_idx, min_widths['default'])
            column_max_width = max_widths.get(col_idx, max_widths['default'])
            
            # Set the width to be between minimum and maximum constraints
            column.width = max(column_min_width, min(calculated_width, column_max_width))
        else:
            # If no content, use the minimum width
            column.width = min_widths.get(col_idx, min_widths['default'])

def copy_row_format(ws, template_row_idx, target_row_idx, max_col=15):
    """
    Copy cell styles (and optionally values) from a template row to a target row.
    Also handles merged cells and maintains their formatting.
    """
    # First, check and unmerge any existing merged cells in the target row
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= target_row_idx <= merged_range.max_row:
            ws.unmerge_cells(str(merged_range))
    
    # Copy format and check for merged cells in template
    template_merged_ranges = []
    for col in range(1, max_col + 1):
        template_cell = ws.cell(row=template_row_idx, column=col)
        target_cell = ws.cell(row=target_row_idx, column=col)
        
        # Copy style and number format
        if hasattr(template_cell, '_style'):
            target_cell._style = copy(template_cell._style)
        if hasattr(template_cell, 'number_format'):
            target_cell.number_format = template_cell.number_format
        if hasattr(template_cell, 'alignment') and template_cell.alignment:
            # Create a new alignment object instead of copying the StyleProxy
            h_align = template_cell.alignment.horizontal if template_cell.alignment.horizontal else 'general'
            v_align = template_cell.alignment.vertical if template_cell.alignment.vertical else 'bottom'
            target_cell.alignment = Alignment(horizontal=h_align, vertical=v_align)
        
        # Check if this cell is part of a merged range in the template
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == template_row_idx and merged_range.min_col <= col <= merged_range.max_col:
                template_merged_ranges.append((
                    merged_range.min_col,
                    merged_range.max_col,
                    merged_range.max_col - merged_range.min_col + 1
                ))
                break
    
    # Recreate merged ranges in the target row
    for min_col, max_col, span in set(template_merged_ranges):
        # Set alignment for all cells in the range before merging
        template_main_cell = ws.cell(row=template_row_idx, column=min_col)
        
        # Get alignment values from template cell
        h_align = 'center'  # Default to center
        v_align = 'center'  # Default to center
        
        if hasattr(template_main_cell, 'alignment') and template_main_cell.alignment:
            if template_main_cell.alignment.horizontal:
                h_align = template_main_cell.alignment.horizontal
            if template_main_cell.alignment.vertical:
                v_align = template_main_cell.alignment.vertical
        
        # Apply alignment to all cells in the range
        for col in range(min_col, max_col + 1):
            target_cell = ws.cell(row=target_row_idx, column=col)
            target_cell.alignment = Alignment(horizontal=h_align, vertical=v_align)
        
        # Now merge the cells
        ws.merge_cells(
            start_row=target_row_idx,
            start_column=min_col,
            end_row=target_row_idx,
            end_column=max_col
        )
        
        # Ensure the merged cell has proper alignment
        merged_cell = ws.cell(row=target_row_idx, column=min_col)
        merged_cell.alignment = Alignment(horizontal=h_align, vertical=v_align)

def copy_merged_and_center(ws, template_ws, template_row_start, template_row_end, target_row_start):
    """
    Copy merged cell structure and center alignment from template header rows to target rows.
    """
    # 1. Copy merged cells
    for m_range in template_ws.merged_cells.ranges:
        if template_row_start <= m_range.min_row <= template_row_end:
            row_offset = target_row_start - template_row_start
            new_min_row = m_range.min_row + row_offset
            new_max_row = m_range.max_row + row_offset
            ws.merge_cells(start_row=new_min_row, start_column=m_range.min_col,
                           end_row=new_max_row, end_column=m_range.max_col)
    # 2. Copy center alignment for header cells
    for row in range(template_row_start, template_row_end + 1):
        for col in range(1, ws.max_column + 1):
            template_cell = template_ws.cell(row=row, column=col)
            if hasattr(template_cell, 'alignment') and template_cell.alignment and template_cell.alignment.horizontal == 'center':
                target_cell = ws.cell(row=target_row_start + (row - template_row_start), column=col)
                # Set center alignment
 
                target_cell.alignment = Alignment(horizontal='center', vertical=template_cell.alignment.vertical)

# Helper function to merge and center header columns
def merge_and_center_header_columns(sheet, start_row, end_row):
    for row in range(start_row, end_row + 1):
        # Set row height to 23.5 for better readability of wrapped text in headers
        sheet.row_dimensions[row].height = 25
        
        # Unmerge any existing merged cells in this row
        for merged_range in list(sheet.merged_cells.ranges):
            if merged_range.min_row <= row <= merged_range.max_row:
                if ((5 <= merged_range.min_col <= 6) or  # SubscriberName (E-F)
                    (7 <= merged_range.min_col <= 9) or  # SystemUser (G-I)
                    (12 <= merged_range.min_col <= 14) or  # DetailsViewedDate (L-N)
                    (15 <= merged_range.min_col <= 17)):  # SearchOutput (O-Q)
                    sheet.unmerge_cells(str(merged_range))
        
        # Merge SubscriberName (E-F)
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        sheet.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        
        # Merge SystemUser (G-I)
        sheet.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        sheet.cell(row=row, column=7).alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge DetailsViewedDate (L-N)
        sheet.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
        sheet.cell(row=row, column=12).alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge SearchOutput (O-Q) with wrap text
        sheet.merge_cells(start_row=row, start_column=15, end_row=row, end_column=17)
        sheet.cell(row=row, column=15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Helper function to merge and center data row columns
def merge_and_center_data_row(sheet, row):
    """Merge and center columns for a specific row.
    Handles SubscriberName (E-F), SystemUser (G-I), DetailsViewedDate (L-N), and SearchOutput (O-Q).
    Also sets row height to 23.5 for better readability of wrapped text.
    """
    # Set row height to 23.5 for better readability of wrapped text
    sheet.row_dimensions[row].height = 25
    
    # Check for existing merged cells in the target row and unmerge them
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row <= row <= merged_range.max_row:
            # Columns for DetailsViewedDate (L-N), SearchOutput (O-Q), SubscriberName (E-F), SystemUser (G-I)
            if (12 <= merged_range.min_col <= 14) or \
                (15 <= merged_range.min_col <= 17) or \
                (5 <= merged_range.min_col <= 6) or \
                (7 <= merged_range.min_col <= 9):
                sheet.unmerge_cells(str(merged_range))
    
    # Set alignment for SubscriberName (E-F)
    for col in range(5, 7):
        cell = sheet.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    
    # Set alignment for SystemUser (G-I)
    for col in range(7, 10):
        cell = sheet.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set alignment for DetailsViewedDate (L-N)
    for col in range(12, 15):
        cell = sheet.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set alignment for SearchOutput (O-Q) with wrap text
    for col in range(15, 18):
        cell = sheet.cell(row=row, column=col)
        # Use top vertical alignment for better readability with wrapped text
        # Center horizontally but align top vertically for multi-line text
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    # Now merge the cells
    # Merge SubscriberName (E-F)
    sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    # Merge SystemUser (G-I)
    sheet.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    # Merge DetailsViewedDate (L-N)
    sheet.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
    # Merge SearchOutput (O-Q)
    sheet.merge_cells(start_row=row, start_column=15, end_row=row, end_column=17)
    
    # Ensure the merged cells have proper alignment
    # SubscriberName
    cell_E = sheet.cell(row=row, column=5)
    cell_E.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # SystemUser
    cell_G = sheet.cell(row=row, column=7)
    cell_G.alignment = Alignment(horizontal='center', vertical='center')
    
    # DetailsViewedDate
    cell_L = sheet.cell(row=row, column=12)
    cell_L.alignment = Alignment(horizontal='center', vertical='center')
    
    # SearchOutput with wrap text
    cell_O = sheet.cell(row=row, column=15)
    # Use top vertical alignment for better readability with wrapped text
    cell_O.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

# Define product rates mapping
PRODUCT_RATES = {
    'Consumer Snap Check': Decimal('500.00'),
    'Consumer Basic Trace': Decimal('170.00'),
    'Consumer Basic Credit': Decimal('170.00'),
    'Consumer Detailed Credit': Decimal('240.00'),
    'X-Score Consumer Detailed Credit': Decimal('500.00'),
    'Commercial Basic Trace': Decimal('275.00'),
    'Commercial Detailed Credit': Decimal('500.00'),
    'Enquiry Report': Decimal('50.00'),
    'Consumer Dud Cheque': Decimal('0.00'),
    'Commercial Dud Cheque': Decimal('0.00'),
    'Director Basic Report': Decimal('0.00'),
    'Director Detailed Report': Decimal('0.00'),
}

def populate_rate_and_amount(ws, start_row, end_row, subscriber_id):
    """
    Populate rate (columns M-O) and calculate amount (columns P-Q) based on product name (column D).
    
    Args:
        ws: Worksheet object
        start_row: First data row (1-based)
        end_row: Last data row (inclusive)
        subscriber_id: The ID/Name of the subscriber to check for custom rates
    """
    
    for row in range(start_row, end_row + 1):
        # Get product name from column D (4th column)
        product_cell = ws.cell(row=row, column=4)
        product_name_original = str(product_cell.value).strip() if product_cell.value else ""
        product_name_cleaned = product_name_original
        
        # Skip empty product names
        if not product_name_cleaned:
            continue
            
        # Use the safe helper function to get the rate
        rate = get_subscriber_product_rate_safe(
            subscriber_id=subscriber_id,
            product_name=product_name_cleaned,
            default_rate_map=PRODUCT_RATES,
            default_rate_key=product_name_cleaned,
            logger=logger
        )
        
        # Populate rate in columns M-O (merged)
        write_to_cell(ws, row, 13, f"₦{rate:,.2f}")  # Column M
        # Calculate amount (rate * 1) since each row represents one search
        amount = rate * Decimal('1.00')
        # Populate amount in columns P-Q (merged)
        write_to_cell(ws, row, 16, f"₦{amount:,.2f}")  # Column P
       

def add_generated_by(ws, username, last_data_row=None):
    """
    Add 'Report Generated by: <username>' two rows below the last data row (or at row 10 if no data), 
    merging O-Q, Trebuchet MS, bold, italic, centered.
    """
    # If no last_data_row provided, default to row 10
    if last_data_row is None:
        last_data_row = 10
    
    # Add two rows below the last data row
    signature_row = last_data_row + 2
    
    # Merge cells O-Q for the signature line
    ws.merge_cells(start_row=signature_row, start_column=15, end_row=signature_row, end_column=17)
    
    # Set the value and formatting for the signature line
    cell = ws.cell(row=signature_row, column=15)
    cell.value = f"Report Generated by: {username}"
    cell.font = openpyxl.styles.Font(name='Trebuchet MS', bold=True, italic=True, color='FF7F7F7F')
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[signature_row].height = 26
    return signature_row

@login_required
def get_subscriber_product_rate(subscriber_name, product_name, default_rate_key, logger=None):
    """Helper function to get subscriber product rate with better error handling."""
    if logger is None:
        logger = logging.getLogger(__name__)
        
    try:
        # Use filter().first() instead of get() to handle multiple records
        custom_rate_obj = SubscriberProductRate.objects.filter(
            subscriber_name__iexact=subscriber_name, 
            product_name__iexact=product_name
        ).first()
        
        if custom_rate_obj and hasattr(custom_rate_obj, 'rate'):
            # Ensure we're returning a Decimal object
            if isinstance(custom_rate_obj.rate, Decimal):
                rate = custom_rate_obj.rate
            else:
                rate = Decimal(str(custom_rate_obj.rate))
            logger.debug(f"Using custom rate for {subscriber_name} - {product_name}: {rate}")
            return rate
        else:
            raise Exception(f"No custom rate found for {product_name}")
    except Exception as e:
        # Get default rate and ensure it's a Decimal
        default_value = ENQUIRY_RATES.get(default_rate_key, Decimal('0.00'))
        if not isinstance(default_value, Decimal):
            rate = Decimal(str(default_value))
        else:
            rate = default_value
        logger.debug(f"Using default rate for {product_name}: {rate}")
        return rate


def get_subscriber_product_rate_safe(subscriber_name, product_name, default_rate_map, default_rate_key, logger=None):
    """Safe helper function to get subscriber product rate with better error handling for multiple records.
    
    Args:
        subscriber_name: The subscriber name to look up
        product_name: The product name to look up
        default_rate_map: Dictionary of default rates (ENQUIRY_RATES or PRODUCT_RATES)
        default_rate_key: Key to use in the default_rate_map
        logger: Optional logger object
    
    Returns:
        Decimal rate value
    """
    if logger is None:
        logger = logging.getLogger(__name__)
        
    try:
        # Use filter().first() instead of get() to avoid MultipleObjectsReturned
        rate_obj = SubscriberProductRate.objects.filter(
            subscriber_name__iexact=subscriber_name, 
            product_name__iexact=product_name
        ).first()
        
        if rate_obj and hasattr(rate_obj, 'rate'):
            # Ensure we're returning a Decimal object
            if isinstance(rate_obj.rate, Decimal):
                rate = rate_obj.rate
            else:
                rate = Decimal(str(rate_obj.rate))
            logger.debug(f"Using custom rate for {subscriber_name} - {product_name}: {rate}")
            return rate
        else:
            # No custom rate found, use default
            default_value = default_rate_map.get(default_rate_key, Decimal('0.00'))
            # Ensure default value is converted to Decimal
            if not isinstance(default_value, Decimal):
                rate = Decimal(str(default_value))
            else:
                rate = default_value
            logger.debug(f"No custom rate found. Using default rate for {product_name}: {rate}")
            return rate
    except Exception as e:
        # Handle any unexpected errors
        logger.error(f"Error retrieving rate for {subscriber_name} - {product_name}: {e}")
        default_value = default_rate_map.get(default_rate_key, Decimal('0.00'))
        # Ensure default value is converted to Decimal
        if not isinstance(default_value, Decimal):
            rate = Decimal(str(default_value))
        else:
            rate = default_value
        logger.debug(f"Using default rate after error: {rate}")
        return rate

        
@login_required
def bulk_report(request):
    """View for generating bulk reports."""
    # Get the current date range (first day of current month to first day of next month)
    today = date.today()
    first_day_of_month = today.replace(day=1)
    
    # Calculate first day of next month
    if today.month == 12:
        first_day_next_month = date(today.year + 1, 1, 1)
    else:
        first_day_next_month = date(today.year, today.month + 1, 1)
    
    # Get unique subscriber names for the dropdown
    subscribers = Usagereport.objects.values_list('SubscriberName', flat=True).distinct().order_by('SubscriberName')
    
    # Initialize report generation tracking
    report_gen = None
    
    # Initial context with date range and subscribers
    context = {
        'start_date': first_day_of_month.strftime('%Y-%m-%d'),
        'end_date': first_day_next_month.strftime('%Y-%m-%d'),
        'subscribers': [{'id': i, 'name': name} for i, name in enumerate(subscribers, 1)],
    }
    
    if request.method == 'POST':
        subscriber_ids = request.POST.getlist('subscribers')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        include_bills = request.POST.get('include_bills') == 'on'
        include_products = request.POST.get('include_products') == 'on'
        
        # Validate that at least one subscriber is selected
        if not subscriber_ids:
            messages.error(request, 'Please select at least one subscriber.')
            return render(request, 'bulkrep/bulk_report.html', context)
            
        # Get the selected subscriber names
        selected_subscribers = [subscribers[int(id)-1] for id in subscriber_ids if int(id) <= len(subscribers)]

        # Create report generation record at the start
        report_gen = ReportGeneration.objects.create(
            user=request.user,
            generator=request.user.username,
            report_type='bulk',
            status='in_progress',
            subscriber_name=f"{len(selected_subscribers)} Subscribers",
            from_date=start_date_str if start_date_str else None,
            to_date=end_date_str if end_date_str else None
        )
        print(f"Started tracking bulk report generation by {request.user.username}")
        
        # Process the form submission

        # Convert date strings to date objects - this is critical for display formatting later
        try:
            # Parse the input date strings
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            # Validate date range
            if start_date > end_date:
                messages.error(request, 'Start date cannot be after end date.')
                return render(request, 'bulkrep/bulk_report.html', context)
                
            # Format the dates for display in the report (DD/MM/YYYY)
            start_date_display = start_date.strftime('%d/%m/%Y')
            end_date_display = end_date.strftime('%d/%m/%Y')
            
        except (ValueError, TypeError) as e:
            messages.error(request, f"Invalid date format: {str(e)}")
            return render(request, 'bulkrep/bulk_report.html', context)

        # Get the selected subscriber names
        subscribers_list = selected_subscribers
        
        # Log the selected subscribers for debugging
        print(f"Selected subscribers: {subscribers_list}")
        
        # Validate that we have subscribers to process
        if not subscribers_list:
            messages.error(request, 'No valid subscribers selected.')
            return render(request, 'bulkrep/bulk_report.html', context)
            
        # Remove any empty strings just in case
        subscribers_list = list(filter(None, set(subscribers_list)))
        
        if not subscribers_list:
            messages.warning(request, f"No subscribers found for the selected criteria between {start_date_display} and {end_date_display}.")
            return render(request, 'bulkrep/bulk_report.html', context)

        # OPTIMIZATION: Fetch all usage data and product rates upfront to eliminate N+1 queries
        print(f"Fetching all usage data for {len(subscribers_list)} subscribers...")
        
        # Fetch all usage data in one query
        all_usage_data = list(Usagereport.objects.filter(
            DetailsViewedDate__gte=start_date,
            DetailsViewedDate__lte=end_date,
            SubscriberName__in=subscribers_list
        ).values(
            'SubscriberName', 'ProductName', 'SystemUser', 'SearchIdentity', 
            'SubscriberEnquiryDate', 'SearchOutput', 'DetailsViewedDate', 
            'ProductInputed'
        ))
        
        # Group usage data by subscriber for instant lookup
        usage_by_subscriber = {}
        for record in all_usage_data:
            subscriber = record['SubscriberName']
            if subscriber not in usage_by_subscriber:
                usage_by_subscriber[subscriber] = []
            usage_by_subscriber[subscriber].append(record)
        
        # Fetch all custom product rates in one query
        print(f"Fetching all custom product rates...")
        all_custom_rates = list(SubscriberProductRate.objects.filter(
            subscriber_name__in=subscribers_list
        ).values('subscriber_name', 'product_name', 'rate'))
        
        # Create lookup dictionary for custom rates: {(subscriber, product): rate}
        custom_rates_lookup = {}
        for rate_record in all_custom_rates:
            key = (rate_record['subscriber_name'].lower(), rate_record['product_name'].lower())
            custom_rates_lookup[key] = Decimal(str(rate_record['rate']))
        
        print(f"Loaded {len(all_usage_data)} usage records and {len(all_custom_rates)} custom rates")
        
        # Helper function to get custom rate with fallback to default
        def get_custom_rate(subscriber_name, product_name, default_rate):
            """Get custom rate for subscriber/product with fallback to default rate."""
            key = (subscriber_name.lower(), product_name.lower())
            return custom_rates_lookup.get(key, Decimal(str(default_rate)))
        
        # Generate reports for all selected subscribers using a zip file
        try:
            template_path = os.path.join(settings.MEDIA_ROOT, 'Templateuse.xlsx')
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                processed_subscribers = []
                
                for subscriber_name in subscribers_list:
                    try:
                        # Log which subscriber we're processing
                        print(f"Processing subscriber: {subscriber_name}")
                        
                        # OPTIMIZED: Use pre-fetched data instead of database queries
                        subscriber_usage_data = usage_by_subscriber.get(subscriber_name, [])
                        
                        if include_bills:
                            # Initialize summary dictionary with all possible keys
                            summary_bills = {
                                'consumer_snap_check': 0,
                                'consumer_basic_trace': 0,
                                'consumer_basic_credit': 0,
                                'consumer_detailed_credit': 0,
                                'xscore_consumer_detailed_credit': 0,
                                'commercial_basic_trace': 0,
                                'commercial_detailed_credit': 0,
                                'enquiry_report': 0,
                                'consumer_dud_cheque': 0,
                                'commercial_dud_cheque': 0,
                                'director_basic_report': 0,
                                'director_detailed_report': 0
                            }
                            
                            # Count each product type using Python filtering (much faster than DB queries)
                            for record in subscriber_usage_data:
                                product_name = record['ProductName']
                                if 'Snap Check' in product_name:
                                    summary_bills['consumer_snap_check'] += 1
                                elif 'Basic Trace' in product_name:
                                    summary_bills['consumer_basic_trace'] += 1
                                elif 'Basic Credit' in product_name:
                                    summary_bills['consumer_basic_credit'] += 1
                                elif 'Detailed Credit' in product_name and 'X-SCore' not in product_name:
                                    summary_bills['consumer_detailed_credit'] += 1
                                elif 'X-SCore Consumer Detailed Credit' in product_name:
                                    summary_bills['xscore_consumer_detailed_credit'] += 1
                                elif 'Commercial Basic Trace' in product_name:
                                    summary_bills['commercial_basic_trace'] += 1
                                elif 'Commercial detailed Credit' in product_name:
                                    summary_bills['commercial_detailed_credit'] += 1
                                elif 'Enquiry Report' in product_name:
                                    summary_bills['enquiry_report'] += 1
                                elif 'Consumer Dud Cheque' in product_name:
                                    summary_bills['consumer_dud_cheque'] += 1
                                elif 'Commercial Dud Cheque' in product_name:
                                    summary_bills['commercial_dud_cheque'] += 1
                                elif 'Director Basic Report' in product_name:
                                    summary_bills['director_basic_report'] += 1
                                elif 'Director Detailed Report' in product_name:
                                    summary_bills['director_detailed_report'] += 1
                        else:
                            summary_bills = {}

                        # OPTIMIZED: Use pre-fetched data for product details
                        if include_products:
                            # Sort the data by ProductName and DetailsViewedDate
                            product_data = sorted(subscriber_usage_data, key=lambda x: (x['ProductName'], x['DetailsViewedDate']))
                            
                            # Group by ProductName
                            product_sections = {}
                            for record in product_data:
                                product_name = record['ProductName']
                                if product_name not in product_sections:
                                    product_sections[product_name] = []
                                product_sections[product_name].append(record)
                        else:
                            product_sections = {}
                            product_data = []

                        if not product_data and include_products:
                            messages.warning(request, f"No data found for subscriber {subscriber_name} between {start_date_display} and {end_date_display}.")
                            continue

                        # Generate Excel report
                        try:
                            template_path = os.path.join(settings.MEDIA_ROOT, 'Templateuse.xlsx')
                            excel_buffer = io.BytesIO()
                            wb = openpyxl.load_workbook(template_path)
                            ws = wb.active
                        except Exception as e:
                            logger.error(f"Error loading Excel template: {str(e)}")
                            messages.error(request, f"Error loading Excel template: {str(e)}")
                            return render(request, 'bulkrep/bulk_report.html', context)
                        try:
                            # Look for "Productname" cell to identify where to put the dynamic product name
                            product_name_cell = None
                            for row in range(30, 35):  # Search rows 30-34
                                for col in range(1, 10):  # Search columns A-I
                                    cell_value = ws.cell(row=row, column=col).value
                                    if cell_value and "product" in str(cell_value).lower():
                                        try:
                                            product_name_cell = (row, col)
                                        except Exception as e:
                                            logger.error(f"Error finding product name cell: {str(e)}")
                                            messages.error(request, f"Error finding product name cell: {str(e)}")
                                            return render(request, 'bulkrep/single_report.html', context)
                                        break
                                if product_name_cell:
                                    break
                            
                            # Write subscriber info directly to cells using the safe method
                            # For row 2, which is merged from H to P, we need to set the value to the first cell in the merged range
                            # First, find the merged range that contains row 2
                            row2_merged_range = None
                            for merged_range in list(ws.merged_cells.ranges):
                                if merged_range.min_row == 2 and merged_range.max_row == 2:
                                    row2_merged_range = merged_range
                                    break
                            
                            # If we found a merged range for row 2, unmerge it, set the value with the required format, and remerge it
                            if row2_merged_range:
                                merged_range_str = str(row2_merged_range)
                                ws.unmerge_cells(merged_range_str)
                                # Set the value to the first cell in the merged range with the required format
                                original_cell = ws.cell(row=2, column=row2_merged_range.min_col)
                                new_content = f"FirstCentral NIGERIA - BILLING DETAILS - {subscriber_name}"
                                original_cell.value = new_content
                                
                                # Remerge the cells
                                ws.merge_cells(merged_range_str)
                            else:
                                # Fallback to the original method if no merged range is found
                                safe_cell_assignment(ws, 2, 5, subscriber_name)  # E2
                            
                            safe_cell_assignment(ws, 5, 4, f"BILLING DETAILS - {subscriber_name}")  # D5
                            
                            # For row 6, which is merged from B to Q, we need to set the value to the first cell in the merged range
                            # First, find the merged range that contains row 6
                            row6_merged_range = None
                            for merged_range in list(ws.merged_cells.ranges):
                                if merged_range.min_row == 6 and merged_range.max_row == 6:
                                    row6_merged_range = merged_range
                                    break
                            
                            # If we found a merged range for row 6, unmerge it, set the value, and remerge it
                            date_range_text = f"REPORT GENERATED FOR RECORDS BETWEEN {start_date_display} and {end_date_display}"
                            if row6_merged_range:
                                merged_range_str = str(row6_merged_range)
                                ws.unmerge_cells(merged_range_str)
                                # Set the value to the first cell in the merged range (column B = 2)
                                ws.cell(row=6, column=2).value = date_range_text
                                # Remerge the cells
                                ws.merge_cells(merged_range_str)
                            else:
                                # Fallback to the original method if no merged range is found
                                safe_cell_assignment(ws, 6, 4, date_range_text)  # D6
                                
                            
                            
                            if include_bills:
                                # Set quantities in column I (existing code)
                                safe_cell_assignment(ws, 12, 9, summary_bills.get('consumer_snap_check', 0) or 0)  # I12
                                safe_cell_assignment(ws, 13, 9, summary_bills.get('consumer_basic_trace', 0) or 0)  # I13
                                safe_cell_assignment(ws, 14, 9, summary_bills.get('consumer_basic_credit', 0) or 0)  # I14
                                safe_cell_assignment(ws, 15, 9, summary_bills.get('consumer_detailed_credit', 0) or 0)  # I15
                                safe_cell_assignment(ws, 16, 9, summary_bills.get('xscore_consumer_detailed_credit', 0) or 0)  # I16
                                safe_cell_assignment(ws, 17, 9, summary_bills.get('commercial_basic_trace', 0) or 0)  # I17
                                safe_cell_assignment(ws, 18, 9, summary_bills.get('commercial_detailed_credit', 0) or 0)  # I18
                                safe_cell_assignment(ws, 20, 9, summary_bills.get('enquiry_report', 0) or 0)  # I20
                                safe_cell_assignment(ws, 22, 9, summary_bills.get('consumer_dud_cheque', 0) or 0)  # I22
                                safe_cell_assignment(ws, 23, 9, summary_bills.get('commercial_dud_cheque', 0) or 0)  # I23
                                safe_cell_assignment(ws, 25, 9, summary_bills.get('director_basic_report', 0) or 0)  # I25
                                safe_cell_assignment(ws, 26, 9, summary_bills.get('director_detailed_report', 0) or 0)  # I26
                                
                                # OPTIMIZED: Set rates in column M (13) using pre-fetched lookup
                                
                                # Consumer Snap Check (row 12)
                                custom_rate = get_custom_rate(subscriber_name, 'Consumer Snap Check', ENQUIRY_RATES['consumer_snap_check'])
                                safe_cell_assignment(ws, 12, 13, f"₦{custom_rate:,.2f}")  # M12
                                
                                # Consumer Basic Trace (row 13)
                                custom_rate = get_custom_rate(subscriber_name, 'Consumer Basic Trace', ENQUIRY_RATES['consumer_basic_trace'])
                                safe_cell_assignment(ws, 13, 13, f"₦{custom_rate:,.2f}")  # M13
                                
                                # Consumer Basic Credit (row 14)
                                custom_rate = get_custom_rate(subscriber_name, 'Consumer Basic Credit', ENQUIRY_RATES['consumer_basic_credit'])
                                safe_cell_assignment(ws, 14, 13, f"₦{custom_rate:,.2f}")  # M14
                                
                                # Consumer Detailed Credit (row 15)
                                custom_rate = get_custom_rate(subscriber_name, 'Consumer Detailed Credit', ENQUIRY_RATES['consumer_detailed_credit'])
                                safe_cell_assignment(ws, 15, 13, f"₦{custom_rate:,.2f}")  # M15
                                
                                # X-Score Consumer Credit (row 16)
                                custom_rate = get_custom_rate(subscriber_name, 'X-Score Consumer Detailed Credit', ENQUIRY_RATES['xscore_consumer_detailed_credit'])
                                safe_cell_assignment(ws, 16, 13, f"₦{custom_rate:,.2f}")  # M16
                                
                                # Commercial Basic Trace (row 17)
                                custom_rate = get_custom_rate(subscriber_name, 'Commercial Basic Trace', ENQUIRY_RATES['commercial_basic_trace'])
                                safe_cell_assignment(ws, 17, 13, f"₦{custom_rate:,.2f}")  # M17
                                
                                # Commercial Detailed Credit (row 18)
                                custom_rate = get_custom_rate(subscriber_name, 'Commercial Detailed Credit', ENQUIRY_RATES['commercial_detailed_credit'])
                                safe_cell_assignment(ws, 18, 13, f"₦{custom_rate:,.2f}")  # M18
                                
                                # Enquiry Report (row 20)
                                custom_rate = get_custom_rate(subscriber_name, 'Enquiry Report', ENQUIRY_RATES['enquiry_report'])
                                safe_cell_assignment(ws, 20, 13, f"₦{custom_rate:,.2f}")  # M20
                                
                                # Consumer Dud Cheque (row 22)
                                custom_rate = get_custom_rate(subscriber_name, 'Consumer Dud Cheque', ENQUIRY_RATES['consumer_dud_cheque'])
                                safe_cell_assignment(ws, 22, 13, f"₦{custom_rate:,.2f}")  # M22
                                
                                # Commercial Dud Cheque (row 23)
                                custom_rate = get_custom_rate(subscriber_name, 'Commercial Dud Cheque', ENQUIRY_RATES['commercial_dud_cheque'])
                                safe_cell_assignment(ws, 23, 13, f"₦{custom_rate:,.2f}")  # M23
                                
                                # Director Basic Report (row 25)
                                custom_rate = get_custom_rate(subscriber_name, 'Director Basic Report', ENQUIRY_RATES['director_basic_report'])
                                safe_cell_assignment(ws, 25, 13, f"₦{custom_rate:,.2f}")  # M25
                                
                                # Director Detailed Report (row 26)
                                custom_rate = get_custom_rate(subscriber_name, 'Director Detailed Report', ENQUIRY_RATES['director_detailed_report'])
                                safe_cell_assignment(ws, 26, 13, f"₦{custom_rate:,.2f}")  # M26
                                            # Calculate and set amounts in column P (16) - merged from P to Q
                                total_amount = 0
                                
                                # Store custom rates in a dictionary for easier reference
                                custom_rates = {}
                                
                                # Calculate amounts (quantity × rate) and populate column P with Naira formatting
                                # Use the optimized custom rate lookup
                                
                                # Consumer Snap Check
                                custom_rate = get_custom_rate(subscriber_name, 'Consumer Snap Check', ENQUIRY_RATES['consumer_snap_check'])
                                consumer_snap_amount = Decimal(str(summary_bills.get('consumer_snap_check', 0) or 0)) * custom_rate
                                safe_cell_assignment(ws, 12, 16, f"₦{consumer_snap_amount:,.2f}")  # P12
                                total_amount += consumer_snap_amount
                                
                                # Consumer Basic Trace
                                custom_rate = get_custom_rate(subscriber_name, 'Consumer Basic Trace', ENQUIRY_RATES['consumer_basic_trace'])
                                consumer_basic_trace_amount = Decimal(str(summary_bills.get('consumer_basic_trace', 0) or 0)) * custom_rate
                                safe_cell_assignment(ws, 13, 16, f"₦{consumer_basic_trace_amount:,.2f}")  # P13
                                total_amount += consumer_basic_trace_amount
                                
                                # Consumer Basic Credit
                                custom_rate = get_custom_rate(subscriber_name, 'Consumer Basic Credit', ENQUIRY_RATES['consumer_basic_credit'])
                                consumer_basic_credit_amount = Decimal(str(summary_bills.get('consumer_basic_credit', 0) or 0)) * custom_rate
                                safe_cell_assignment(ws, 14, 16, f"₦{consumer_basic_credit_amount:,.2f}")  # P14
                                total_amount += consumer_basic_credit_amount
                                
                                # Consumer Detailed Credit
                                custom_rate = get_custom_rate(subscriber_name, 'Consumer Detailed Credit', ENQUIRY_RATES['consumer_detailed_credit'])
                                consumer_detailed_credit_amount = Decimal(str(summary_bills.get('consumer_detailed_credit', 0) or 0)) * custom_rate
                                safe_cell_assignment(ws, 15, 16, f"₦{consumer_detailed_credit_amount:,.2f}")  # P15
                                total_amount += consumer_detailed_credit_amount
                                
                                # X-Score Consumer Credit
                                custom_rate = get_custom_rate(subscriber_name, 'X-Score Consumer Detailed Credit', ENQUIRY_RATES['xscore_consumer_detailed_credit'])
                                xscore_consumer_amount = Decimal(str(summary_bills.get('xscore_consumer_detailed_credit', 0) or 0)) * custom_rate
                                safe_cell_assignment(ws, 16, 16, f"₦{xscore_consumer_amount:,.2f}")  # P16
                                total_amount += xscore_consumer_amount
                                
                                # Commercial Basic Trace
                                custom_rate = get_custom_rate(subscriber_name, 'Commercial Basic Trace', ENQUIRY_RATES['commercial_basic_trace'])
                                commercial_basic_trace_amount = Decimal(str(summary_bills.get('commercial_basic_trace', 0) or 0)) * custom_rate
                                safe_cell_assignment(ws, 17, 16, f"₦{commercial_basic_trace_amount:,.2f}")  # P17
                                total_amount += commercial_basic_trace_amount
                                
                                # Commercial Detailed Credit
                                custom_rate = get_custom_rate(subscriber_name, 'Commercial Detailed Credit', ENQUIRY_RATES['commercial_detailed_credit'])
                                commercial_detailed_credit_amount = Decimal(str(summary_bills.get('commercial_detailed_credit', 0) or 0)) * custom_rate
                                safe_cell_assignment(ws, 18, 16, f"₦{commercial_detailed_credit_amount:,.2f}")  # P18
                                total_amount += commercial_detailed_credit_amount
                                
                                # Enquiry Report
                                custom_rates['enquiry_report'] = get_custom_rate(subscriber_name, 'Enquiry Report', ENQUIRY_RATES['enquiry_report'])
                                logger.debug(f"Using rate for {subscriber_name} - Enquiry Report: {custom_rates['enquiry_report']}")
                                
                                enquiry_report_amount = Decimal(str(summary_bills.get('enquiry_report', 0) or 0)) * custom_rates['enquiry_report']
                                safe_cell_assignment(ws, 20, 16, f"₦{enquiry_report_amount:,.2f}")  # P20
                                total_amount += enquiry_report_amount
                                
                                # Consumer Dud Cheque
                                custom_rates['consumer_dud_cheque'] = get_custom_rate(subscriber_name, 'Consumer Dud Cheque', ENQUIRY_RATES['consumer_dud_cheque'])
                                logger.debug(f"Using rate for {subscriber_name} - Consumer Dud Cheque: {custom_rates['consumer_dud_cheque']}")
                                
                                consumer_dud_amount = Decimal(str(summary_bills.get('consumer_dud_cheque', 0) or 0)) * custom_rates['consumer_dud_cheque']
                                safe_cell_assignment(ws, 22, 16, f"₦{consumer_dud_amount:,.2f}")  # P22
                                total_amount += consumer_dud_amount
                                
                                # Commercial Dud Cheque
                                custom_rates['commercial_dud_cheque'] = get_custom_rate(subscriber_name, 'Commercial Dud Cheque', ENQUIRY_RATES['commercial_dud_cheque'])
                                logger.debug(f"Using rate for {subscriber_name} - Commercial Dud Cheque: {custom_rates['commercial_dud_cheque']}")
                                
                                commercial_dud_amount = Decimal(str(summary_bills.get('commercial_dud_cheque', 0) or 0)) * custom_rates['commercial_dud_cheque']
                                safe_cell_assignment(ws, 23, 16, f"₦{commercial_dud_amount:,.2f}")  # P23
                                total_amount += commercial_dud_amount
                                
                                # Director Basic Report
                                custom_rates['director_basic_report'] = get_custom_rate(subscriber_name, 'Director Basic Report', ENQUIRY_RATES['director_basic_report'])
                                logger.debug(f"Using rate for {subscriber_name} - Director Basic Report: {custom_rates['director_basic_report']}")
                                
                                director_basic_amount = Decimal(str(summary_bills.get('director_basic_report', 0) or 0)) * custom_rates['director_basic_report']
                                safe_cell_assignment(ws, 25, 16, f"₦{director_basic_amount:,.2f}")  # P25
                                total_amount += director_basic_amount
                                
                                # Director Detailed Report
                                custom_rates['director_detailed_report'] = get_custom_rate(subscriber_name, 'Director Detailed Report', ENQUIRY_RATES['director_detailed_report'])
                                logger.debug(f"Using rate for {subscriber_name} - Director Detailed Report: {custom_rates['director_detailed_report']}")
                                
                                director_detailed_amount = Decimal(str(summary_bills.get('director_detailed_report', 0) or 0)) * custom_rates['director_detailed_report']
                                safe_cell_assignment(ws, 26, 16, f"₦{director_detailed_amount:,.2f}")  # P26
                                total_amount += director_detailed_amount
                                
                                # Set total amount with Naira formatting
                                safe_cell_assignment(ws, 28, 16, f"₦{total_amount:,.2f}")  # P28 for total
                                
                                # Calculate 7.5% VAT amount
                                vat_amount = total_amount * Decimal(str(0.075))
                                safe_cell_assignment(ws, 29, 16, f"₦{vat_amount:,.2f}")  # P29 for VAT amount
                                
                                # Calculate amount due (total + VAT)
                                amount_due = total_amount + vat_amount
                                safe_cell_assignment(ws, 30, 16, f"₦{amount_due:,.2f}")  # P30 for amount due
                                

                            
                            if include_products:
                                start_row_offset = 36  # Initial start row for product sections on Sheet1
                                current_sheet = ws
                                sheet2 = wb["Sheet2"] if "Sheet2" in wb.sheetnames else None
                                
                                # Sort product_sections by product_name
                                sorted_product_names = sorted(product_sections.keys())
                                
                                # First, find the original product name cell in the template (rows 32-35)
                                product_name_cell = None
                                for row in range(32, 36):  # Rows 32-35 (inclusive)
                                    for col in range(1, 16):  # Assuming columns A-O are important
                                        cell_value = ws.cell(row=row, column=col).value
                                        if cell_value and "product" in str(cell_value).lower():
                                            product_name_cell = (row, col)
                                            break
                                    if product_name_cell:
                                        break
                                        
                                # Save the template header structure (rows 32-35) for subsequent products
                                header_template = []
                                header_rows = (32, 35)  # Range of rows to copy for the header template
                                for row in range(header_rows[0], header_rows[1] + 1):
                                    row_data = []
                                    for col in range(1, 16):  # Assuming columns A-O are important
                                        cell = ws.cell(row=row, column=col)
                                        # Store cell value and position only - we'll copy styles directly later
                                        cell_info = {
                                            'value': cell.value,
                                            'position': (row, col)
                                        }
                                        # Check if cell is part of merged range
                                        for m_range in ws.merged_cells.ranges:
                                            if (row, col) == (m_range.min_row, m_range.min_col):
                                                cell_info['merged'] = (m_range.max_row - m_range.min_row + 1, 
                                                                    m_range.max_col - m_range.min_col + 1)
                                                break
                                        row_data.append(cell_info)
                                    header_template.append(row_data)
                                
                                # Set data row start - will be used for first product                
                                data_start_row = 36  # Initial start for data rows after template header
                                lastProduct = ""
                                
                                # For each product, use appropriate header section
                                for product_idx, product_name in enumerate(sorted_product_names):
                                    product_records = product_sections[product_name]
                                    
                                    if product_idx == 0:
                                        # For first product, use the existing template header
                                        if product_name_cell:
                                            row, col = product_name_cell
                                            # Replace "Product Name" with actual first product name in template
                                            safe_cell_assignment(ws, row, col, product_name)
                                        current_sheet = ws
                                        current_row_offset = data_start_row  # Start data at row 36
                                        serial_number_base = data_start_row - 1
                                    else:
                                        # Add space between different products (extra spacing)
                                        current_row_offset += 4  # Add significant spacing between products
                                            
                                        # Create new header for subsequent products
                                        header_start_row = current_row_offset
                                        
                                        # Clone the header section for this product
                                        for template_row_idx, template_row in enumerate(header_template):
                                            target_row = header_start_row + template_row_idx
                                        
                                            # Ensure we're not past row limits
                                            if current_sheet == ws and target_row > 1000000 and sheet2:
                                                current_sheet = sheet2
                                                # Reset for Sheet2
                                                header_start_row = 13
                                                target_row = header_start_row + template_row_idx
                                        
                                            # Unmerge any existing merged cells in the target area
                                            for m_range in list(current_sheet.merged_cells.ranges):
                                                if m_range.min_row <= target_row <= m_range.max_row:
                                                    current_sheet.unmerge_cells(
                                                        start_row=m_range.min_row, 
                                                        start_column=m_range.min_col,
                                                        end_row=m_range.max_row, 
                                                        end_column=m_range.max_col
                                                    )
                                        
                                            # Copy each cell from the template to the target area
                                            for col_idx, cell_info in enumerate(template_row):
                                                target_col = col_idx + 1
                                                target_cell = current_sheet.cell(row=target_row, column=target_col)
                                                
                                                # Copy styles directly from the original cell
                                                original_row, original_col = cell_info['position']
                                                original_cell = ws.cell(row=original_row, column=original_col)
                                                
                                                # Copy cell format using openpyxl's built-in method
                                                target_cell._style = copy(original_cell._style)
                                                
                                                # Set value using safe_cell_assignment to handle merged cells properly
                                                if template_row_idx == product_name_cell[0] - header_rows[0] and \
                                                col_idx == product_name_cell[1] - 1:
                                                    # For the product name cell, use the safe assignment method
                                                    safe_cell_assignment(current_sheet, target_row, target_col, product_name)
                                                elif cell_info['value'] is not None:
                                                    # For other cells with values, use the safe assignment method
                                                    safe_cell_assignment(current_sheet, target_row, target_col, cell_info['value'])
                                                
                                                # Recreate merged cells
                                                if 'merged' in cell_info:
                                                    rows, cols = cell_info['merged']
                                                    current_sheet.merge_cells(
                                                        start_row=target_row, 
                                                        start_column=target_col,
                                                        end_row=target_row + rows - 1, 
                                                        end_column=target_col + cols - 1
                                                    )
                                        
                                        # Update data start row to be after this new header
                                        current_row_offset = header_start_row + (header_rows[1] - header_rows[0] + 1)
                                        serial_number_base = current_row_offset - 1
                                        
                                        # Add header for the data section
                                        safe_cell_assignment(current_sheet, current_row_offset - 1, 4, "Unique Tracking Number")

                                    # Process data records for this product
                                    for record_idx, record in enumerate(product_records):
                                        current_row = current_row_offset + record_idx
                                        
                                        # Switch to Sheet2 when reaching max row (like VBA)
                                        if current_row > 1000000 and sheet2 and current_sheet != sheet2:
                                            current_sheet = sheet2
                                            current_row_offset = 13  # Reset row offset for Sheet2 as per VBA
                                            serial_number_base = 12  # Reset serial number base for Sheet2
                                            current_row = current_row_offset + record_idx  # Recalculate current_row for Sheet2
                                        
                                        # Copy row format from template data row
                                        template_data_row = 36
                                        max_col = 17  # Extend to include all columns that need formatting
                                        copy_row_format(current_sheet, template_data_row, current_row, max_col)
                                        
                                        # Now assign values as before
                                        safe_cell_assignment(current_sheet, current_row, 2, current_row - serial_number_base)  # Serial Number
                                        safe_cell_assignment(current_sheet, current_row, 3, "")  # Branch ID
                                        safe_cell_assignment(current_sheet, current_row, 4, "")  # Unique Tracking Number column left blank
                                        safe_cell_assignment(current_sheet, current_row, 5, record['SubscriberName'])
                                        safe_cell_assignment(current_sheet, current_row, 7, record['SystemUser'] if record['SystemUser'] else "")
                                        # Format SubscriberEnquiryDate
                                        subscriber_enquiry_date = record['SubscriberEnquiryDate']
                                        if subscriber_enquiry_date:
                                            if isinstance(subscriber_enquiry_date, str):
                                                date_str = subscriber_enquiry_date
                                            else:
                                                # Convert to date object first to remove any time component
                                                if hasattr(subscriber_enquiry_date, 'date'):
                                                    date_only = subscriber_enquiry_date.date()
                                                else:
                                                    date_only = subscriber_enquiry_date
                                                date_str = date_only.strftime('%Y-%m-%d')
                                            # Set cell format to text BEFORE assignment to prevent Excel from adding time component
                                            current_sheet.cell(row=current_row, column=10).number_format = '@'
                                            safe_cell_assignment(current_sheet, current_row, 10, date_str)
                                            # Ensure the format stays as text after assignment
                                            current_sheet.cell(row=current_row, column=10).number_format = '@'
                                        else:
                                            safe_cell_assignment(current_sheet, current_row, 10, "")
                                        safe_cell_assignment(current_sheet, current_row, 11, record['ProductName'])
                                        # Format DetailsViewedDate to match SubscriberEnquiryDate format
                                        details_viewed_date = record['DetailsViewedDate']
                                        if details_viewed_date:
                                            if isinstance(details_viewed_date, str):
                                                date_str = details_viewed_date
                                            else:
                                                # Convert to date object first to remove any time component
                                                if hasattr(details_viewed_date, 'date'):
                                                    date_only = details_viewed_date.date()
                                                else:
                                                    date_only = details_viewed_date
                                                date_str = date_only.strftime('%Y-%m-%d')
                                            # Set cell format to text BEFORE assignment to prevent Excel from adding time component
                                            current_sheet.cell(row=current_row, column=12).number_format = '@'
                                            safe_cell_assignment(current_sheet, current_row, 12, date_str)
                                            # Ensure the format stays as text after assignment
                                            current_sheet.cell(row=current_row, column=12).number_format = '@'
                                        else:
                                            safe_cell_assignment(current_sheet, current_row, 12, "")
                                        safe_cell_assignment(current_sheet, current_row, 15, record['SearchOutput'] if record['SearchOutput'] else "")
                                        
                                        # Apply merge and center formatting to this data row
                                        merge_and_center_data_row(current_sheet, current_row)
                                    
                                    # Move offset to after this product's data for next product
                                    current_row_offset += len(product_records)
                                    
                                    # Set lastProduct for the next iteration
                                    lastProduct = product_name

                            # Auto-size columns for better readability
                            for sheet in wb.worksheets:
                                auto_size_columns(sheet)

                            add_generated_by(ws, request.user.username, current_row_offset - 1)

                        except Exception as e:
                            logger.error(f"Error generating report data: {str(e)}")
                            messages.error(request, f"Error generating report data: {str(e)}")
                            return render(request, 'bulkrep/bulk_report.html', context)
                        
                        # Save to buffer
                        wb.save(excel_buffer)
                        excel_buffer.seek(0)
                        
                        # Prepare filename using the subscriber name
                        month_year = start_date.strftime('%B%Y')
                        clean_subscriber = clean_filename(subscriber_name)
                        filename = f"{clean_subscriber}_{month_year}.xlsx"
                        
                        # Add to zip file
                        zip_file.writestr(filename, excel_buffer.getvalue())
                        
                        # Add to processed subscribers list
                        processed_subscribers.append(subscriber_name)
                        
                    except Exception as e:
                        print(f"Error processing subscriber {subscriber_name}: {str(e)}")
                        error_msg = f"Skipped report for {subscriber_name} due to error: {str(e)}"
                        messages.warning(request, error_msg)
                        if 'report_gen' in locals() and report_gen:
                            report_gen.status = 'failed'
                            report_gen.error_message = error_msg[:500]  # Truncate to fit in the field
                            report_gen.completed_at = timezone.now()
                            report_gen.save()
                        continue
            
            # Return the zip file if any reports were generated
            if processed_subscribers:
                zip_buffer.seek(0)
                month_year = start_date.strftime('%B%Y')
                zip_filename = f"all_subscriber_reports_{month_year}_{uuid.uuid4().hex[:8]}.zip"
                bulk_reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'bulk')
                os.makedirs(bulk_reports_dir, exist_ok=True)
                zip_path = os.path.join(bulk_reports_dir, zip_filename)
                with open(zip_path, 'wb') as f:
                    f.write(zip_buffer.read())
                download_url = settings.MEDIA_URL + f'reports/bulk/{zip_filename}'
                
                # Update report generation status to success
                if 'report_gen' in locals() and report_gen:
                    report_gen.status = 'success'
                    report_gen.completed_at = timezone.now()
                    report_gen.save()
                
                success_msg = f"Successfully generated reports for {len(processed_subscribers)} out of {len(subscribers_list)} subscribers for the period {start_date_display} to {end_date_display}."
                messages.success(request, success_msg)
                return render(request, 'bulkrep/download_ready.html', {
                    'download_url': download_url
                })
            else:
                messages.warning(request, "No reports were generated. Please check the data or try different criteria.")
                return render(request, 'bulkrep/bulk_report.html', context)
            
        except Exception as e:
            messages.error(request, f"Error generating bulk reports: {str(e)}")
            return render(request, 'bulkrep/bulk_report.html', context)
    
    # For GET requests, render the bulk report form
    return render(request, 'bulkrep/bulk_report.html', context)


# Dashboard Views


@login_required
@user_passes_test(lambda u: u.is_superuser)
def dashboard(request):
    """Main dashboard view"""
    context = {
        'title': '📊 Usage Analytics Dashboard'
    }
    return render(request, 'bulkrep/dashboard.html', context)

# MODIFIED: Refactored dashboard_api for simplicity and correctness
# In views.py, replace your entire dashboard_api function with this one.

# In views.py, replace the entire dashboard_api function

@login_required
def dashboard_api(request):
    """API endpoint for dashboard data with caching and consistent filtering."""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return JsonResponse({'error': 'Authentication required or insufficient permissions'}, status=403)

    try:
        # --- Date Handling ---
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            today = timezone.now().date()
            start_date = today.replace(day=1)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            today = timezone.now().date()
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1)

        # --- Filter Handling ---
        subscriber_filter = request.GET.get('subscriber_filter')
        if subscriber_filter and subscriber_filter.lower().strip() in ['all', 'null', '']:
            subscriber_filter = None
        
        # --- Special View Handling ---
        if request.GET.get('three_month_view', 'false').lower() == 'true':
            data = {'three_month_usage': get_three_month_rolling_usage(subscriber_filter)}
            return JsonResponse(data)

        # --- Cache Key Generation ---
        cache_params = {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'subscriber_filter': subscriber_filter or 'all',
        }
        cache_key_string = '|'.join([f"{k}:{v}" for k, v in sorted(cache_params.items())])
        cache_key = f"dashboard_api_v6_{hashlib.md5(cache_key_string.encode()).hexdigest()}" # Changed to v6 to be safe

        cached_data = cache.get(cache_key)
        if cached_data:
            return JsonResponse(cached_data)

        # --- UNIFIED DATA PAYLOAD ---
        data = {
            'total_subscribers': get_total_subscribers(start_date, end_date, subscriber_filter),
            'total_usage_entries': get_total_usage_entries(start_date, end_date, subscriber_filter),
            'top_subscriber': get_top_subscriber(start_date, end_date, subscriber_filter),
            'top_subscribers': get_top_subscribers_by_usage_filtered(start_date, end_date, subscriber_filter, limit=10),
            'top_products': get_top_products_by_frequency_filtered(start_date, end_date, subscriber_filter=subscriber_filter, limit=10),
            'all_products_by_frequency': get_all_products_by_frequency(start_date, end_date, subscriber_filter),
            'all_subscribers_by_usage': get_all_subscribers_by_usage(start_date, end_date, subscriber_filter),
            'key_subscribers_list': KEY_SUBSCRIBERS, # This line ensures the key list is always sent
            'usage_trends': get_usage_trends_filtered(start_date, end_date, subscriber_filter=subscriber_filter),
            'new_subscribers': get_new_subscribers_trend_filtered(start_date, end_date, subscriber_filter=subscriber_filter),
            'retention_rate': get_retention_rate(start_date, end_date, subscriber_filter),
            'highest_product_by_transaction': get_highest_product_by_transaction(start_date, end_date, subscriber_filter),
            'unique_products': get_unique_products(subscriber_filter),
            'unique_subscribers': get_unique_subscribers(),
            'daily_comparison': get_daily_comparison(subscriber_filter),
            'churn_data': get_churn_data_filtered(start_date, end_date, subscriber_filter=subscriber_filter),
        }

        cache.set(cache_key, data, 300)
        return JsonResponse(data)

    except Exception as e:
        import traceback
        logger.error(f"Dashboard API error: {traceback.format_exc()}")
        return JsonResponse({'error': 'An error occurred while loading dashboard data.'}, status=500)


# MODIFIED: This function is now correct
def get_total_subscribers(start_date, end_date, subscriber_filter=None):
    """Get total number of unique subscribers for date range, filtered by subscriber if provided"""
    query = Usagereport.objects.filter(
        DetailsViewedDate__range=[start_date, end_date]
    )
    if subscriber_filter == 'key_subscribers':
        query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        return query.values('SubscriberName').distinct().count()
    elif subscriber_filter and subscriber_filter != 'all':
        # Check if the single subscriber has any usage in the period
        return 1 if query.filter(SubscriberName=subscriber_filter).exists() else 0
    else:
        return query.values('SubscriberName').distinct().count()

# MODIFIED: This function is now correct
def get_total_usage_entries(start_date, end_date, subscriber_filter=None):
    """Get total usage entries for date range, filtered by subscriber if provided"""
    query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])

    if subscriber_filter == 'key_subscribers':
        query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
    elif subscriber_filter and subscriber_filter != 'all':
        query = query.filter(SubscriberName=subscriber_filter)

    return query.count()

# MODIFIED: Added subscriber_filter
def get_top_subscriber(start_date, end_date, subscriber_filter=None):
    """Get the top subscriber by usage count within the filtered group."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        elif subscriber_filter and subscriber_filter != 'all':
            return subscriber_filter # If filtering for one, they are the top one
        
        top_subscriber = query.values('SubscriberName').annotate(
            usage_count=Count('SearchIdentity')
        ).order_by('-usage_count').first()
        
        return top_subscriber['SubscriberName'] if top_subscriber else 'N/A'
    except Exception as e:
        logger.error(f"Error getting top subscriber: {str(e)}")
        return 'N/A'

# MODIFIED: Added subscriber_filter
def get_unique_products(subscriber_filter=None):
    """Get list of unique product names for filter dropdown, optionally filtered."""
    try:
        query = Usagereport.objects
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)
            
        products = query.values_list('ProductName', flat=True).distinct().order_by('ProductName')
        # return [product for product in products if product]
        return list(products)
    except Exception as e:
        logger.error(f"Error getting unique products: {str(e)}")
        return []

def get_unique_subscribers():
    """Get list of ALL unique subscriber names for the global filter dropdown."""
    try:
        subscribers = Usagereport.objects.values_list('SubscriberName', flat=True).distinct().order_by('SubscriberName')
        return [subscriber for subscriber in subscribers if subscriber]
    except Exception as e:
        logger.error(f"Error getting unique subscribers: {str(e)}")
        return []

# MODIFIED: Simplified to one function and corrected logic
def get_top_subscribers_by_usage_filtered(start_date, end_date, subscriber_filter=None, limit=10):
    """Get top subscribers by usage volume with consistent filtering."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)
        
        top_subscribers = query.values('SubscriberName').annotate(
            usage_count=Count('SearchIdentity')
        ).order_by('-usage_count')
        
        if limit:
            top_subscribers = top_subscribers[:limit]
            
        return list(top_subscribers)
    except Exception as e:
        logger.error(f"Error getting filtered subscribers: {str(e)}")
        return []

# In views.py
def get_all_products_by_frequency(start_date, end_date, subscriber_filter=None):
    """Gets ALL products by frequency with consistent filtering, without a limit."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )

        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)

        all_products = query.values('ProductName').annotate(
            frequency=Count('ProductName')
        ).order_by('-frequency')

        return list(all_products)
    except Exception as e:
        logger.error(f"Error getting all products by frequency: {str(e)}")
        return []

# MODIFIED: Simplified to one function and corrected logic
def get_top_products_by_frequency_filtered(start_date, end_date, product_filter=None, subscriber_filter=None, limit=25):
    """Get top products by frequency with consistent filtering."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)

        if product_filter and product_filter != 'all':
            query = query.filter(ProductName__icontains=product_filter)
        
        top_products = query.values('ProductName').annotate(
            frequency=Count('ProductName')
        ).order_by('-frequency')[:limit]
        
        return list(top_products)
    except Exception as e:
        logger.error(f"Error getting filtered top products: {str(e)}")
        return []

# MODIFIED: Added subscriber_filter and corrected period logic
def get_churn_data_filtered(start_date, end_date, churn_days=None, subscriber_filter=None):
    """Get churn data with optional day filtering - HIGHLY OPTIMIZED"""
    try:
        # Date period logic
        duration_days = (end_date - start_date).days
        previous_end = start_date - timedelta(days=1)
        previous_start = previous_end - timedelta(days=duration_days)

        # Base queries
        previous_period_query = Usagereport.objects.filter(DetailsViewedDate__range=[previous_start, previous_end])
        current_period_query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])

        # Apply subscriber filter if relevant (for key subscribers)
        if subscriber_filter == 'key_subscribers':
            previous_period_query = previous_period_query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
            current_period_query = current_period_query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        # Note: Churn for a single subscriber is not meaningful, so we don't handle that case.

        previous_subscribers = set(previous_period_query.values_list('SubscriberName', flat=True).distinct())
        current_subscribers = set(current_period_query.values_list('SubscriberName', flat=True).distinct())

        churned_subscribers_set = previous_subscribers - current_subscribers
        
        churned_count = len(churned_subscribers_set)
        previous_subscribers_count = len(previous_subscribers)
        current_subscribers_count = len(current_subscribers)
        
        churn_rate = (churned_count / previous_subscribers_count * 100) if previous_subscribers_count else 0
        
        return {
            'churned_count': churned_count,
            'churn_rate': round(churn_rate, 2),
            'previous_subscribers': previous_subscribers_count,
            'current_subscribers': current_subscribers_count
        }
    except Exception as e:
        logger.error(f"Error getting filtered churn data: {str(e)}")
        return {'churned_count': 0, 'churn_rate': 0, 'previous_subscribers': 0, 'current_subscribers': 0}

# MODIFIED: Added subscriber_filter
def get_retention_rate(start_date, end_date, subscriber_filter=None):
    """Calculate subscriber retention rate with filtering."""
    try:
        duration_days = (end_date - start_date).days
        previous_end = start_date - timedelta(days=1)
        previous_start = previous_end - timedelta(days=duration_days)

        previous_period_query = Usagereport.objects.filter(DetailsViewedDate__range=[previous_start, previous_end])
        current_period_query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])

        if subscriber_filter == 'key_subscribers':
            previous_period_query = previous_period_query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
            current_period_query = current_period_query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        
        previous_subscribers = set(previous_period_query.values_list('SubscriberName', flat=True).distinct())
        current_subscribers = set(current_period_query.values_list('SubscriberName', flat=True).distinct())
        
        retained_subscribers = previous_subscribers.intersection(current_subscribers)
        retention_rate = len(retained_subscribers) / len(previous_subscribers) * 100 if previous_subscribers else 0
        
        return {
            'retention_rate': round(retention_rate, 2),
            'retained_count': len(retained_subscribers),
            'previous_count': len(previous_subscribers)
        }
    except Exception as e:
        logger.error(f"Error getting retention rate: {e}")
        return {'retention_rate': 0}

# MODIFIED: This function is now correct
def get_usage_trends_filtered(start_date, end_date, usage_trends_days=None, subscriber_filter=None):
    """Get usage trends with optional day filtering and subscriber filtering"""
    try:
        if usage_trends_days and usage_trends_days.isdigit():
            days = int(usage_trends_days)
            trends_start = end_date - timedelta(days=days)
        else:
            trends_start = start_date
        
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[trends_start, end_date]
        )
        
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)
        
        return list(query.annotate(
            date=Cast('DetailsViewedDate', DateField())
        ).values('date').annotate(
            count=Count('ProductName') # Corrected field
        ).order_by('date'))
    except Exception as e:
        logger.error(f"Error getting filtered usage trends: {str(e)}")
        return []


def get_all_subscribers_by_usage(start_date, end_date, subscriber_filter=None):
    """Gets ALL subscribers by usage with consistent filtering, without a limit."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )

        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)

        all_subscribers = query.values('SubscriberName').annotate(
            usage_count=Count('SubscriberName')  # Counting SubscriberName to include all records
        ).order_by('-usage_count')

        return list(all_subscribers)
    except Exception as e:
        logger.error(f"Error getting all subscribers by usage: {str(e)}")
        return []
def get_new_subscribers_trend_filtered(start_date, end_date, new_subscribers_days=None, subscriber_filter=None):
    """Get new subscribers trend with optional filtering."""
    try:
        if new_subscribers_days and new_subscribers_days.isdigit():
            days = int(new_subscribers_days)
            trends_start = end_date - timedelta(days=days)
        else:
            trends_start = start_date
        
        # Base query to find the first usage date for each subscriber
        first_usage_query = Usagereport.objects.values('SubscriberName').annotate(first_usage=Min('DetailsViewedDate'))

        # Apply subscriber filter if provided
        if subscriber_filter == 'key_subscribers':
            first_usage_query = first_usage_query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        elif subscriber_filter and subscriber_filter != 'all':
             first_usage_query = first_usage_query.filter(SubscriberName=subscriber_filter)

        # Filter the results to the desired date range
        first_usage_dates = first_usage_query.filter(first_usage__range=[trends_start, end_date])
        
        # Group by date in memory
        trend_data = defaultdict(int)
        for item in first_usage_dates:
            first_usage_date = item['first_usage']
            if isinstance(first_usage_date, datetime):
                first_usage_date = first_usage_date.date()
            if isinstance(first_usage_date, date):
                trend_data[first_usage_date] += 1
        
        # Create complete date range with zero counts for missing dates
        new_subscribers_by_date = []
        current_date = trends_start
        while current_date <= end_date:
            new_subscribers_by_date.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'new_subscribers': trend_data.get(current_date, 0)
            })
            current_date += timedelta(days=1)
        
        return new_subscribers_by_date
    except Exception as e:
        logger.error(f"Error getting filtered new subscribers trend: {str(e)}")
        return []

# MODIFIED: Added subscriber_filter
def get_highest_product_by_transaction(start_date, end_date, subscriber_filter=None):
    """Get the product with the highest number of transactions within the filtered group."""
    try:
        query = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date]
        )
        if subscriber_filter == 'key_subscribers':
            query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        elif subscriber_filter and subscriber_filter != 'all':
            query = query.filter(SubscriberName=subscriber_filter)

        # In views.py
        top_product = query.values('ProductName').annotate(
            transaction_count=Count('SearchIdentity') # Corrected field
        ).order_by('-transaction_count').first()
        
        return top_product['ProductName'] if top_product else 'N/A'
    except Exception as e:
        logger.error(f"Error getting highest product by transaction: {str(e)}")
        return 'N/A'

# MODIFIED: Added subscriber_filter and corrected logic
def get_three_month_rolling_usage(subscriber_filter=None):
    """Get usage data for the current month and two previous months, broken down by month."""
    try:
        today = timezone.now().date()
        months_data = []
        
        for i in range(3): # Loop for 3 months
            # Determine the start and end of the month
            first_day_of_month = (today.replace(day=1) - timedelta(days=i*30)).replace(day=1)
            last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

            query = Usagereport.objects.filter(DetailsViewedDate__range=[first_day_of_month, last_day_of_month])
            
            if subscriber_filter == 'key_subscribers':
                query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
            elif subscriber_filter and subscriber_filter != 'all':
                query = query.filter(SubscriberName=subscriber_filter)
            
            usage_count = query.count()
            month_name = first_day_of_month.strftime("%B")
            
            months_data.append({
                'month': f"{month_name} {first_day_of_month.year}",
                'usage_count': usage_count,
                'month_short': month_name[:3],
                'year': first_day_of_month.year
            })
        
        return list(reversed(months_data)) # Show oldest to newest
        
    except Exception as e:
        logger.error(f"Error getting 3-month rolling usage: {str(e)}")
        return []

# MODIFIED: Corrected filter logic
def get_daily_comparison(subscriber_filter=None):
    """Returns usage counts for yesterday and the same day of the previous month."""
    yesterday = (timezone.now() - timedelta(days=1)).date()
    # Safely get the same day last month
    try:
        prev_month_day = yesterday.replace(month=yesterday.month - 1)
    except ValueError: # Handles month-end cases like March 31 -> February
        prev_month_day = yesterday - timedelta(days=28)
        prev_month_day = prev_month_day.replace(day=yesterday.day)

    query_yesterday = Usagereport.objects.filter(DetailsViewedDate=yesterday)
    query_prev = Usagereport.objects.filter(DetailsViewedDate=prev_month_day)

    if subscriber_filter == 'key_subscribers':
        query_yesterday = query_yesterday.filter(SubscriberName__in=KEY_SUBSCRIBERS)
        query_prev = query_prev.filter(SubscriberName__in=KEY_SUBSCRIBERS)
    elif subscriber_filter and subscriber_filter != 'all':
        query_yesterday = query_yesterday.filter(SubscriberName=subscriber_filter)
        query_prev = query_prev.filter(SubscriberName=subscriber_filter)

    return {
        'yesterday': {
            'date': yesterday.strftime('%Y-%m-%d'),
            'count': query_yesterday.count(),
        },
        'previous_month_same_day': {
            'date': prev_month_day.strftime('%Y-%m-%d'),
            'count': query_prev.count(),
        }
    }

# --- The remaining functions like download_churned_subscribers etc. are not part of the dashboard API flow ---
# --- and do not need modification for this specific task. They remain as they are. ---

def get_churned_subscribers_list(start_date, end_date, churn_days=None):
    """Get list of churned subscriber names for download"""
    try:
        # Use custom days if provided, otherwise use default logic
        if churn_days and churn_days.isdigit():
            days = int(churn_days)
            analysis_start = end_date - timedelta(days=days)
            previous_start = analysis_start - timedelta(days=days)
            previous_end = analysis_start
        else:
            # Use original logic
            analysis_start = start_date
            previous_start = start_date - timedelta(days=(end_date - start_date).days)
            previous_end = start_date
        
        # Get previous subscribers
        previous_subscribers = set(Usagereport.objects.filter(
            DetailsViewedDate__range=[previous_start, previous_end]
        ).values_list('SubscriberName', flat=True).distinct())
        
        # Get current subscribers
        current_subscribers = set(Usagereport.objects.filter(
            DetailsViewedDate__range=[analysis_start, end_date]
        ).values_list('SubscriberName', flat=True).distinct())
        
        # Calculate churned subscribers
        churned_subscribers = previous_subscribers - current_subscribers
        
        # Return sorted list
        return sorted(list(churned_subscribers))
        
    except Exception as e:
        logger.error(f"Error getting churned subscribers list: {str(e)}")
        return []

# In views.py, add this new function at the end of the file.

# @login_required
# In views.py, replace the previous download_top_subscribers_csv function

# @login_required
# def download_top_subscribers_csv(request):
#     """
#     Generates and serves a TSV (Tab-Separated Values) file with a detailed breakdown of usage 
#     for the top 10 subscribers, respecting all active dashboard filters.
#     """
#     try:
#         # Get and validate filters from the request
#         start_date_str = request.GET.get('start_date')
#         end_date_str = request.GET.get('end_date')
#         subscriber_filter = request.GET.get('subscriber_filter')

#         today = timezone.now().date()
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
        
#         if end_date_str:
#             end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
#         else:
#             end_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

#         # Build the base queryset with date and subscriber filters
#         base_query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])

#         if subscriber_filter == 'key_subscribers':
#             base_query = base_query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
#         elif subscriber_filter and subscriber_filter != 'all':
#             base_query = base_query.filter(SubscriberName=subscriber_filter)

#         # First, find the names of the top 10 subscribers within the filtered group
#         top_subscribers_names = list(base_query.values('SubscriberName')
#                                     .annotate(total_usage=Count('SearchIdentity'))
#                                     .order_by('-total_usage')[:10]
#                                     .values_list('SubscriberName', flat=True))

#         if not top_subscribers_names:
#             return HttpResponse("No data available for the selected filters.", content_type="text/plain")

#         # Now, get the detailed product breakdown for ONLY those top subscribers
#         detailed_data = base_query.filter(SubscriberName__in=top_subscribers_names) \
#                                   .values('SubscriberName', 'ProductName') \
#                                   .annotate(usage_count=Count('SearchIdentity')) \
#                                   .order_by('SubscriberName', '-usage_count')

#         # Generate the TSV (tab-separated) response
#         response = HttpResponse(content_type='text/tab-separated-values')
#         filename = f"top_subscribers_usage_detail_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.tsv"
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'

#         writer = csv.writer(response, delimiter='\t')
#         writer.writerow(['Subscriber Name', 'Product Name', 'Usage Count'])

#         for row in detailed_data:
#             writer.writerow([
#                 row['SubscriberName'],
#                 row['ProductName'],
#                 row['usage_count']
#             ])

#         return response

#     except Exception as e:
#         logger.error(f"Error downloading top subscribers TSV: {e}")
#         return HttpResponse("An error occurred while generating the report.", status=500)
# In views.py, replace the entire download_top_subscribers_csv function.

# @login_required
# def download_top_subscribers_csv(request):
#     """
#     Generates a TSV file for top subscriber usage.
#     - If a specific subscriber is selected, it downloads their detailed usage.
#     - Otherwise, it downloads the detailed usage for the top 10 subscribers in the current filter group.
#     """
#     try:
#         # Get filters from the request
#         start_date_str = request.GET.get('start_date')
#         end_date_str = request.GET.get('end_date')
#         global_subscriber_filter = request.GET.get('subscriber_filter')
#         selected_subscriber = request.GET.get('selected_subscriber', None)

#         # Handle date ranges
#         today = timezone.now().date()
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
#         if end_date_str:
#             end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
#         else:
#             end_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

#         # Build the base queryset with the date filter
#         base_query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])
#         filename_part = ""

#         if selected_subscriber:
#             # Case 1: A specific subscriber was selected from the dropdown.
#             # Filter for that single subscriber's detailed product usage.
#             detailed_data = base_query.filter(SubscriberName=selected_subscriber) \
#                                       .values('SubscriberName', 'ProductName') \
#                                       .annotate(usage_count=Count('SearchIdentity')) \
#                                       .order_by('-usage_count')
#             filename_part = clean_filename(selected_subscriber)
#         else:
#             # Case 2: No specific subscriber selected, so get the top 10 from the global filter group.
            
#             # Apply the global filter to the base query
#             if global_subscriber_filter == 'key_subscribers':
#                 base_query = base_query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
#                 filename_part = "Key_Subscribers_Top_10"
#             elif global_subscriber_filter and global_subscriber_filter != 'all':
#                 base_query = base_query.filter(SubscriberName=global_subscriber_filter)
#                 filename_part = clean_filename(global_subscriber_filter)
#             else:
#                 filename_part = "All_Subscribers_Top_10"

#             # Get the names of the top 10 subscribers from the filtered group
#             top_subscribers_names = list(base_query.values('SubscriberName')
#                                         .annotate(total_usage=Count('SearchIdentity'))
#                                         .order_by('-total_usage')[:10]
#                                         .values_list('SubscriberName', flat=True))

#             if not top_subscribers_names:
#                 return HttpResponse("No data available for the selected filters.", content_type="text/plain")

#             # Get the detailed product usage for ONLY those top 10 subscribers
#             detailed_data = base_query.filter(SubscriberName__in=top_subscribers_names) \
#                                       .values('SubscriberName', 'ProductName') \
#                                       .annotate(usage_count=Count('SearchIdentity')) \
#                                       .order_by('SubscriberName', '-usage_count')
        
#         # Generate the TSV response
#         response = HttpResponse(content_type='text/tab-separated-values')
#         filename = f"{filename_part}_usage_detail_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.tsv"
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'

#         writer = csv.writer(response, delimiter='\t')
#         writer.writerow(['Subscriber Name', 'Product Name', 'Usage Count'])

#         for row in detailed_data:
#             writer.writerow([
#                 row['SubscriberName'],
#                 row['ProductName'],
#                 row['usage_count']
#             ])

#         return response

#     except Exception as e:
#         logger.error(f"Error downloading top subscribers TSV: {e}")
#         return HttpResponse("An error occurred while generating the report.", status=500)

# In views.py, replace the download_top_subscribers_csv function

@login_required
def download_top_subscribers_csv(request):
    """
    Generates a TSV file for top subscriber usage. This version is optimized to
    prevent database timeouts by ensuring querysets are constructed efficiently.
    """
    try:
        # 1. Get and parse request parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        global_subscriber_filter = request.GET.get('subscriber_filter')
        selected_subscriber = request.GET.get('selected_subscriber', None)

        today = timezone.now().date()
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        filename_part = ""
        
        # 2. Determine the query based on parameters
        if selected_subscriber:
            # Case 1: A specific subscriber was selected.
            filename_part = clean_filename(selected_subscriber)
            detailed_data = Usagereport.objects.filter(
                DetailsViewedDate__range=[start_date, end_date],
                SubscriberName=selected_subscriber
            ).values('SubscriberName', 'ProductName') \
             .annotate(usage_count=Count('SearchIdentity')) \
             .order_by('-usage_count')

        else:
            # Case 2: No specific subscriber, so find the Top 10 for the global filter group.
            query = Usagereport.objects.filter(DetailsViewedDate__range=[start_date, end_date])
            
            if global_subscriber_filter == 'key_subscribers':
                query = query.filter(SubscriberName__in=KEY_SUBSCRIBERS)
                filename_part = "Key_Subscribers_Top_10"
            elif global_subscriber_filter and global_subscriber_filter != 'all':
                query = query.filter(SubscriberName=global_subscriber_filter)
                filename_part = clean_filename(global_subscriber_filter)
            else:
                filename_part = "All_Subscribers_Top_10"

            # This is the potentially slow query. The indexes will speed it up.
            top_subscribers_names = list(query.values('SubscriberName')
                                        .annotate(total_usage=Count('SearchIdentity'))
                                        .order_by('-total_usage')[:10]
                                        .values_list('SubscriberName', flat=True))

            if not top_subscribers_names:
                return HttpResponse("No data available for the selected filters.", content_type="text/plain")

            detailed_data = query.filter(SubscriberName__in=top_subscribers_names) \
                                 .values('SubscriberName', 'ProductName') \
                                 .annotate(usage_count=Count('SearchIdentity')) \
                                 .order_by('SubscriberName', '-usage_count')

        # 3. Generate and stream the TSV response
        response = HttpResponse(content_type='text/tab-separated-values')
        filename = f"{filename_part}_usage_detail_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.tsv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response, delimiter='\t')
        writer.writerow(['Subscriber Name', 'Product Name', 'Usage Count'])

        for row in detailed_data:
            writer.writerow([row['SubscriberName'], row['ProductName'], row['usage_count']])

        return response

    except Exception as e:
        logger.error(f"Error downloading top subscribers TSV: {e}")
        return HttpResponse("An error occurred while generating the report.", status=500)


def download_churned_subscribers(request):
    """Download churned subscribers as CSV file"""

    
    try:
        # Get date range from request
        start_date_str = request.GET.get('start_date', None)
        end_date_str = request.GET.get('end_date', None)
        churn_days = request.GET.get('churn_days', None)
        
        # Parse dates or use defaults
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            # Default to first day of current month
            today = timezone.now().date()
            start_date = today.replace(day=1)
            
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            # Default to first day of next month
            today = timezone.now().date()
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1)
        
        # Get churned subscribers list
        churned_subscribers = get_churned_subscribers_list(start_date, end_date, churn_days)
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="churned_subscribers_{start_date}_{end_date}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Subscriber Name'])  # Header
        
        # Write subscriber names
        for subscriber in churned_subscribers:
            writer.writerow([subscriber])
        
        return response
        
    except Exception as e:
        logger.error(f"Error downloading churned subscribers: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)


def download_new_subscribers(request):
    """Download new subscribers as TXT file with subscriber names and join dates"""
    from django.http import HttpResponse
    from datetime import datetime
    
    try:
        # Get date range from request
        start_date_str = request.GET.get('start_date', None)
        end_date_str = request.GET.get('end_date', None)
        
        # Parse dates or use defaults
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            # Default to 30 days ago
            today = timezone.now().date()
            start_date = today - timedelta(days=30)
            
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            # Default to today
            end_date = timezone.now().date()
        
        # Get new subscribers details with names and join dates
        new_subscribers_details = get_new_subscribers_details(start_date, end_date)
        
        # Create TXT response
        response = HttpResponse(content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="new_subscribers_{start_date}_{end_date}.txt"'
        
        # Write header
        content = f"New Subscribers Report\n"
        content += f"Date Range: {start_date} to {end_date}\n"
        content += f"Total New Subscribers: {len(new_subscribers_details)}\n\n"
        content += f"{'Subscriber Name':<50} {'Date Joined':<15}\n"
        content += f"{'-' * 50} {'-' * 15}\n"
        
        # Write subscriber details
        for subscriber in new_subscribers_details:
            content += f"{subscriber['name']:<50} {subscriber['date_joined']:<15}\n"
        
        response.write(content)
        return response
        
    except Exception as e:
        logger.error(f"Error downloading new subscribers: {str(e)}")
        return HttpResponse(f"Error: {str(e)}", status=500)


def new_subscribers_trend_api(request):
    """API endpoint for new subscribers trend data with custom date range"""
    try:
        # Get date range from request
        start_date_str = request.GET.get('start_date', None)
        end_date_str = request.GET.get('end_date', None)
        
        if not start_date_str or not end_date_str:
            return JsonResponse({'error': 'Both start_date and end_date are required'}, status=400)
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Validate date range
        if start_date > end_date:
            return JsonResponse({'error': 'Start date cannot be later than end date'}, status=400)
        
        # Get new subscribers data
        new_subscribers_data = get_new_subscribers_trend_optimized(start_date, end_date)
        
        return JsonResponse({
            'new_subscribers': new_subscribers_data,
            'start_date': start_date_str,
            'end_date': end_date_str
        })
        
    except ValueError as e:
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
    except Exception as e:
        logger.error(f"Error in new subscribers trend API: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


def usage_trends_api(request):
    """API endpoint for usage trends data with custom date range"""
    try:
        # Get date range from request
        start_date_str = request.GET.get('start_date', None)
        end_date_str = request.GET.get('end_date', None)
        
        if not start_date_str or not end_date_str:
            return JsonResponse({'error': 'Both start_date and end_date are required'}, status=400)
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Validate date range
        if start_date > end_date:
            return JsonResponse({'error': 'Start date cannot be later than end date'}, status=400)
        
        # Get subscriber filter if provided
        subscriber_filter = request.GET.get('subscriber_filter', 'all')
        if subscriber_filter == 'all':
            subscriber_filter = None
        
        # Get usage trends data
        usage_trends_data = get_usage_trends_filtered(start_date, end_date, None, subscriber_filter)
        
        return JsonResponse({
            'usage_trends': usage_trends_data,
            'start_date': start_date_str,
            'end_date': end_date_str
        })
        
    except ValueError as e:
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
    except Exception as e:
        logger.error(f"Error in usage trends API: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)

def get_subscriber_product_rate(subscriber, product_key):
    """
    Get the subscriber-specific rate for a product with fallback to ENQUIRY_RATES.
    
    Args:
        subscriber (str): The subscriber name
        product_key (str): The product key in ENQUIRY_RATES format (with underscores)
    
    Returns:
        Decimal: The rate for the subscriber-product combination
    """
    try:
        # Map product keys to actual product names in database
        product_variations = [
            product_key.replace('_', ' '),
            product_key.replace('_', '-'),
            product_key.title().replace('_', ' '),
        ]
        
        # Try to find a subscriber-specific rate
        for variation in product_variations:
            try:
                # Use filter().first() to handle duplicate records gracefully
                subscriber_rate = SubscriberProductRate.objects.filter(
                    subscriber_name=subscriber,
                    product_name__icontains=variation
                ).first()
                if subscriber_rate:
                    return subscriber_rate.rate
            except Exception:
                continue
        
        # Fall back to ENQUIRY_RATES if no specific rate exists
        return ENQUIRY_RATES.get(product_key, Decimal('0.00'))
    except Exception as e:
        logger.error(f"Error getting subscriber product rate: {str(e)}")
        return ENQUIRY_RATES.get(product_key, Decimal('0.00'))

def get_all_subscriber_product_rate():
    """
    Gets all product subscriber rates in a single query and groups them in memory.
    """
    try:
        # Use iterator for memory efficiency on large datasets
        all_rates_data = SubscriberProductRate.objects.values_list(
            'subscriber_name', 'product_name', 'rate'
        ).iterator(chunk_size=2000)

        # Group rates by subscriber for fast lookups
        grouped_rates = defaultdict(dict)
        for subscriber_name, product_name, rate in all_rates_data:
            # Store as {subscriber: {product: rate}}
            grouped_rates[subscriber_name][product_name.lower()] = rate
            
        return grouped_rates
        
    except Exception as e:
        logger.error(f"Error getting all subscriber product rates: {str(e)}")
        return {} # Return an empty dictionary on error


def get_top_products_by_subscriber_filtered(start_date, end_date, subscriber_name):
    """Show only products used by selected subscriber"""
    try:
        from django.db.models import Count
        
        if not subscriber_name:
            return []
        
        # Get products used by the specific subscriber
        products = Usagereport.objects.filter(
            DetailsViewedDate__range=[start_date, end_date],
            SubscriberName=subscriber_name
        ).values('ProductName').annotate(
            usage_count=Count('ProductName')
        ).order_by('-usage_count')[:10]
        
        result = []
        for product in products:
            result.append({
                'product': product['ProductName'],
                'count': product['usage_count']
            })
        
        return result
    except Exception as e:
        # logger.error(f"Error getting top products by subscriber: {str(e)}")
        return []


def get_new_subscribers_details(start_date, end_date):
    """Get detailed list of new subscribers with names and join dates"""
    try:
        from datetime import datetime, date

        # Ensure start_date and end_date are consistently date objects
        if isinstance(start_date, datetime):
            start_date = start_date.date()
        elif isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

        if isinstance(end_date, datetime):
            end_date = end_date.date()
        elif isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Get first usage date for each subscriber with their names
        new_subscribers = Usagereport.objects.values('SubscriberName').annotate(
            first_usage=Min('DetailsViewedDate')
        ).filter(
            first_usage__range=[start_date, end_date]
        ).order_by('first_usage', 'SubscriberName')

        # Format the results
        subscribers_details = []
        for item in new_subscribers:
            first_usage = item['first_usage']
            if first_usage is None:
                continue

            # Ensure first_usage is a date object
            if isinstance(first_usage, datetime):
                first_usage = first_usage.date()
            elif isinstance(first_usage, str):
                first_usage = datetime.strptime(first_usage.split(' ')[0], '%Y-%m-%d').date()
            
            if isinstance(first_usage, date):
                subscribers_details.append({
                    'name': item['SubscriberName'],
                    'date_joined': first_usage.strftime('%Y-%m-%d')
                })
        
        return subscribers_details
    except Exception as e:
        logger.error(f"Error getting new subscribers details: {str(e)}")
        return []


def get_new_subscribers_trend_optimized(start_date, end_date):
    """Get new subscribers trend - OPTIMIZED version without day filtering"""
    try:
        

        # Ensure start_date and end_date are consistently date objects
        if isinstance(start_date, datetime):
            start_date = start_date.date()
        elif isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

        if isinstance(end_date, datetime):
            end_date = end_date.date()
        elif isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Single query to get first usage date for each subscriber
        first_usage_dates = Usagereport.objects.values('SubscriberName').annotate(
            first_usage=Min('DetailsViewedDate')
        ).filter(
            first_usage__range=[start_date, end_date]
        )

        # Group by date
        trend_data = {}
        for item in first_usage_dates:
            first_usage = item['first_usage']
            if first_usage is None:
                continue

            # Defensively ensure first_usage is a date object before calling strftime
            if isinstance(first_usage, datetime):
                first_usage = first_usage.date()
            elif isinstance(first_usage, str):
                # If the database returns a string, convert it. This handles formats like 'YYYY-MM-DD' or 'YYYY-MM-DD HH:MM:SS'
                first_usage = datetime.strptime(first_usage.split(' ')[0], '%Y-%m-%d').date()
            
            # Now we can safely call strftime
            if isinstance(first_usage, date):
                date_key = first_usage.strftime('%Y-%m-%d')
                trend_data[date_key] = trend_data.get(date_key, 0) + 1
        
        # Create complete date range with zero counts for missing dates
        new_subscribers_by_date = []
        current_date = start_date
        
        # The redundant checks inside the loop have been removed for cleanliness and efficiency.
        while current_date <= end_date:
            date_key = current_date.strftime('%Y-%m-%d')
            new_subscribers_by_date.append({
                'date': date_key,
                'new_subscribers': trend_data.get(date_key, 0)
            })
            current_date += timedelta(days=1)
        
        return new_subscribers_by_date
    except Exception as e:
        logger.error(f"Error getting new subscribers trend optimized: {str(e)}")
        return []

        