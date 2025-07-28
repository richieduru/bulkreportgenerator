# Template Optimization System for Excel Report Generation
# This module creates optimized template variants from the institutional template
# while preserving all original formatting and compliance requirements

import os
import shutil
from pathlib import Path
from django.conf import settings
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import logging
from typing import Dict, List, Optional, Tuple
from copy import copy
from openpyxl.cell import MergedCell

logger = logging.getLogger(__name__)

class TemplateOptimizer:
    """
    Creates optimized template variants from the institutional base template.
    Maintains all original formatting while pre-configuring layouts for different scenarios.
    """
    
    def __init__(self):
        self.base_template_path = os.path.join(settings.MEDIA_ROOT, 'Templateuse.xlsx')
        self.templates_dir = os.path.join(settings.MEDIA_ROOT, 'optimized_templates')
        self.ensure_templates_directory()
        
        # Template variants configuration
        self.template_variants = {
            'bills_only': {
                'filename': 'template_bills_only.xlsx',
                'description': 'Pre-configured for reports with billing summary only',
                'has_products': False,
                'max_bill_rows': 20
            },
            'products_light': {
                'filename': 'template_products_light.xlsx', 
                'description': 'Optimized for reports with up to 100 product records',
                'has_products': True,
                'max_product_rows': 100,
                'product_sections': 3
            },
            'products_heavy': {
                'filename': 'template_products_heavy.xlsx',
                'description': 'Optimized for reports with 100+ product records',
                'has_products': True,
                'max_product_rows': 1000,
                'product_sections': 10
            },
            'bulk_single': {
                'filename': 'template_bulk_single.xlsx',
                'description': 'Optimized for single subscriber in bulk operations',
                'has_products': True,
                'simplified_formatting': True
            }
        }
    
    def ensure_templates_directory(self):
        """Create optimized templates directory if it doesn't exist."""
        Path(self.templates_dir).mkdir(parents=True, exist_ok=True)
    
    def create_all_template_variants(self) -> Dict[str, str]:
        """
        Create all optimized template variants from the base template.
        Returns dictionary mapping variant names to file paths.
        """
        if not os.path.exists(self.base_template_path):
            raise FileNotFoundError(f"Base template not found: {self.base_template_path}")
        
        created_templates = {}
        
        for variant_name, config in self.template_variants.items():
            try:
                template_path = self.create_template_variant(variant_name, config)
                created_templates[variant_name] = template_path
                logger.info(f"Created optimized template: {variant_name}")
            except Exception as e:
                logger.error(f"Failed to create template variant {variant_name}: {str(e)}")
                
        return created_templates
    
    def create_template_variant(self, variant_name: str, config: Dict) -> str:
        """
        Create a specific template variant based on configuration.
        """
        # Load base template
        wb = openpyxl.load_workbook(self.base_template_path)
        ws = wb.active
        
        # Apply variant-specific optimizations
        if variant_name == 'bills_only':
            self._optimize_bills_only_template(ws, config)
        elif variant_name in ['products_light', 'products_heavy']:
            self._optimize_products_template(ws, config)
        elif variant_name == 'bulk_single':
            self._optimize_bulk_single_template(ws, config)
        
        # Save optimized template
        output_path = os.path.join(self.templates_dir, config['filename'])
        wb.save(output_path)
        wb.close()
        
        return output_path
    
    def _optimize_bills_only_template(self, ws, config: Dict):
        """
        Optimize template for bills-only reports.
        Pre-formats billing section and removes product-related areas.
        """
        # Pre-expand billing rows with formatting
        billing_start_row = 12  # Based on template analysis
        max_bill_rows = config.get('max_bill_rows', 20)
        
        # Copy billing row format to additional rows
        template_bill_row = 12
        for i in range(1, max_bill_rows):
            target_row = billing_start_row + i
            self._copy_row_formatting(ws, template_bill_row, target_row)
        
        # Pre-merge cells for billing section
        for row in range(billing_start_row, billing_start_row + max_bill_rows):
            # Pre-merge common billing columns
            self._apply_billing_merges(ws, row)
        
        # Clear product section area (rows 30+) but keep basic structure
        self._clear_product_section(ws, keep_headers=False)
    
    def _optimize_products_template(self, ws, config: Dict):
        """
        Optimize template for product-heavy reports.
        Pre-creates multiple product sections with formatting.
        """
        max_rows = config.get('max_product_rows', 100)
        sections = config.get('product_sections', 3)
        
        # Identify product template area (rows 32-36 based on analysis)
        product_header_start = 32
        product_header_end = 35
        product_data_start = 36
        
        # Pre-create multiple product sections
        current_row = product_data_start
        rows_per_section = max_rows // sections
        
        for section in range(sections):
            # Copy product header for each section
            if section > 0:
                self._copy_product_header(ws, product_header_start, product_header_end, current_row)
                current_row += (product_header_end - product_header_start + 1)
            
            # Pre-format data rows for this section
            for i in range(rows_per_section):
                self._copy_row_formatting(ws, product_data_start, current_row + i)
                self._apply_product_data_merges(ws, current_row + i)
            
            current_row += rows_per_section + 2  # Add spacing between sections
    
    def _optimize_bulk_single_template(self, ws, config: Dict):
        """
        Optimize template for bulk processing of single subscribers.
        Simplified formatting for faster processing.
        """
        # Pre-apply common merges to reduce runtime operations
        self._pre_apply_common_merges(ws)
        
        # Simplify auto-sizing by pre-setting optimal column widths
        self._set_optimal_column_widths(ws)
        
        # Pre-format common data areas
        self._pre_format_data_areas(ws)
    
    def _is_merged_cell(self, ws, row: int, col: int) -> bool:
        """
        Check if a cell is part of a merged range.
        """
        cell = ws.cell(row=row, column=col)
        return isinstance(cell, MergedCell)
    
    def _copy_row_formatting(self, ws, source_row: int, target_row: int):
        """
        Copy all formatting from source row to target row.
        """
        for col in range(1, 18):  # Columns A-Q
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)
            
            # Skip merged cells to avoid read-only errors
            if self._is_merged_cell(ws, source_row, col):
                continue
            
            # Copy style
            if hasattr(source_cell, '_style'):
                target_cell._style = copy(source_cell._style)
            
            # Copy number format
            target_cell.number_format = source_cell.number_format
            
            # Copy alignment
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
    
    def _copy_product_header(self, ws, start_row: int, end_row: int, target_start: int):
        """
        Copy product header section to a new location.
        """
        for row_offset in range(end_row - start_row + 1):
            source_row = start_row + row_offset
            target_row = target_start + row_offset
            
            for col in range(1, 18):
                source_cell = ws.cell(row=source_row, column=col)
                target_cell = ws.cell(row=target_row, column=col)
                
                # Skip merged cells to avoid read-only errors
                if self._is_merged_cell(ws, source_row, col):
                    continue
                
                # Copy value and formatting
                if not isinstance(source_cell, MergedCell):
                    target_cell.value = source_cell.value
                    
                if hasattr(source_cell, '_style'):
                    target_cell._style = copy(source_cell._style)
                target_cell.number_format = source_cell.number_format
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
    
    def _apply_billing_merges(self, ws, row: int):
        """
        Apply standard billing row merges.
        """
        # Based on template analysis, merge specific column ranges
        merge_ranges = [
            (1, 7),   # A-G for bill description
            (8, 11),  # H-K for quantity
            (12, 14), # L-N for rate
            (15, 17)  # O-Q for amount
        ]
        
        for start_col, end_col in merge_ranges:
            if start_col != end_col:
                ws.merge_cells(
                    start_row=row, start_column=start_col,
                    end_row=row, end_column=end_col
                )
                # Set center alignment
                ws.cell(row=row, column=start_col).alignment = Alignment(
                    horizontal='center', vertical='center'
                )
    
    def _apply_product_data_merges(self, ws, row: int):
        """
        Apply standard product data row merges.
        """
        # Based on template analysis
        merge_ranges = [
            (5, 6),   # E-F for SubscriberName
            (7, 9),   # G-I for SystemUser
            (12, 14), # L-N for DetailsViewedDate
            (15, 17)  # O-Q for SearchOutput
        ]
        
        for start_col, end_col in merge_ranges:
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col
            )
            # Set appropriate alignment
            alignment = Alignment(horizontal='center', vertical='center')
            if start_col == 15:  # SearchOutput column - use wrap text
                alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            
            ws.cell(row=row, column=start_col).alignment = alignment
    
    def _clear_product_section(self, ws, keep_headers: bool = True):
        """
        Clear product section while optionally keeping headers.
        """
        start_row = 36 if not keep_headers else 37
        end_row = 100  # Clear reasonable range
        
        for row in range(start_row, end_row):
            for col in range(1, 18):
                cell = ws.cell(row=row, column=col)
                # Skip merged cells to avoid read-only errors
                if not self._is_merged_cell(ws, row, col) and not isinstance(cell, MergedCell):
                    cell.value = None
    
    def _pre_apply_common_merges(self, ws):
        """
        Pre-apply all common merge operations to reduce runtime overhead.
        """
        # Apply merges for header rows
        header_rows = [2, 6, 8, 9]  # Based on template analysis
        for row in header_rows:
            # Apply standard header merges
            ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=15)
    
    def _set_optimal_column_widths(self, ws):
        """
        Pre-set optimal column widths based on content analysis.
        """
        optimal_widths = {
            1: 8,   # A
            2: 10,  # B
            3: 8,   # C
            4: 15,  # D
            5: 12,  # E
            6: 12,  # F
            7: 10,  # G
            8: 10,  # H
            9: 10,  # I
            10: 15, # J
            11: 20, # K
            12: 12, # L
            13: 12, # M
            14: 12, # N
            15: 25, # O
            16: 25, # P
            17: 25  # Q
        }
        
        for col_num, width in optimal_widths.items():
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width
    
    def _pre_format_data_areas(self, ws):
        """
        Pre-format common data entry areas.
        """
        # Set row heights for better readability
        for row in range(36, 136):  # Data rows
            ws.row_dimensions[row].height = 25
        
        # Pre-apply text format to date columns
        for row in range(36, 136):
            ws.cell(row=row, column=10).number_format = '@'  # SubscriberEnquiryDate
            ws.cell(row=row, column=12).number_format = '@'  # DetailsViewedDate
    
    def get_optimal_template(self, has_products: bool, product_count: int = 0, 
                           is_bulk: bool = False) -> str:
        """
        Return the optimal template path based on report characteristics.
        """
        if not has_products:
            return os.path.join(self.templates_dir, 'template_bills_only.xlsx')
        
        if is_bulk:
            return os.path.join(self.templates_dir, 'template_bulk_single.xlsx')
        
        if product_count <= 100:
            return os.path.join(self.templates_dir, 'template_products_light.xlsx')
        else:
            return os.path.join(self.templates_dir, 'template_products_heavy.xlsx')
    
    def template_exists(self, variant_name: str) -> bool:
        """
        Check if a template variant exists.
        """
        if variant_name not in self.template_variants:
            return False
        
        template_path = os.path.join(
            self.templates_dir, 
            self.template_variants[variant_name]['filename']
        )
        return os.path.exists(template_path)
    
    def refresh_templates(self) -> Dict[str, str]:
        """
        Refresh all template variants from the base template.
        Useful when the institutional template is updated.
        """
        logger.info("Refreshing optimized templates from base template")
        
        # Remove existing optimized templates
        if os.path.exists(self.templates_dir):
            shutil.rmtree(self.templates_dir)
        
        # Recreate directory and templates
        self.ensure_templates_directory()
        return self.create_all_template_variants()


# Template Manager for integration with existing views
class OptimizedTemplateManager:
    """
    Manager class to integrate optimized templates with existing report generation.
    """
    
    def __init__(self):
        self.optimizer = TemplateOptimizer()
        self._ensure_templates_ready()
    
    def _ensure_templates_ready(self):
        """
        Ensure all template variants are available.
        """
        missing_templates = []
        for variant_name in self.optimizer.template_variants.keys():
            if not self.optimizer.template_exists(variant_name):
                missing_templates.append(variant_name)
        
        if missing_templates:
            logger.info(f"Creating missing template variants: {missing_templates}")
            self.optimizer.create_all_template_variants()
    
    def get_optimized_workbook(self, has_products: bool = True, 
                             product_count: int = 0, is_bulk: bool = False):
        """
        Get an optimized workbook instance based on report requirements.
        This replaces the create_workbook_from_template() function.
        """
        template_path = self.optimizer.get_optimal_template(
            has_products=has_products,
            product_count=product_count,
            is_bulk=is_bulk
        )
        
        if not os.path.exists(template_path):
            logger.warning(f"Optimized template not found: {template_path}")
            # Fallback to base template
            template_path = self.optimizer.base_template_path
        
        return openpyxl.load_workbook(template_path)
    
    def refresh_all_templates(self):
        """
        Refresh all optimized templates from the institutional base template.
        """
        return self.optimizer.refresh_templates()


# Global instance for use in views
template_manager = OptimizedTemplateManager()